from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse

from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Query
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from pathlib import Path
from sqlalchemy import select, func, outerjoin, delete, exists
from sqlalchemy.orm import selectinload # <-- ADD THIS
from typing import List, Dict, Optional, Any, Union, Tuple

import openpyxl
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from starlette.middleware.cors import CORSMiddleware
from contextlib import asynccontextmanager
# --- ADD THESE IMPORTS ---
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi import Depends
from sqlalchemy import select # Make sure to add this import at the top of the file


# Import our new modules
import models
import database
# --- END OF IMPORTS TO ADD ---
import os
import logging
import json
import re
import io
import asyncio

from pydantic import BaseModel, Field
from typing import List, Dict, Optional, Any, Union
import uuid
from datetime import datetime, timezone
import pandas as pd
from pandas import NaT
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import KMeans
from sklearn.metrics.pairwise import cosine_similarity
import openai
import requests
from collections import defaultdict, Counter

# OpenAI setup
openai.api_key = os.environ.get('OPENAI_API_KEY')
LOCAL_LLM_ENDPOINT = os.environ.get('LOCAL_LLM_ENDPOINT')

# Create the main app without a prefix
# --- ADD THIS ENTIRE BLOCK OF CODE ---

@asynccontextmanager
async def lifespan(app: FastAPI):
    """
    An asynchronous context manager to handle startup and shutdown events.
    """
    # This part of the code will run ONCE when the server starts.
    logger.info("Server is starting up...")
    
    yield  # The server runs after this 'yield'
    
    # This part of the code will run ONCE when the server shuts down.
    # We no longer need to close the database client here, as each
    # session is managed by the get_db dependency.
    logger.info("Server is shutting down...")


# Now, define the 'app' variable that uses the function above
app = FastAPI(title="Tally Statement Processor", version="1.0.0", lifespan=lifespan)
# --- END OF BLOCK TO ADD ---
# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Pydantic models for Client data
class ClientBase(BaseModel):
    """Defines the common attributes for a client."""
    name: str

class ClientCreate(ClientBase):
    """The model used when CREATING a new client via the API."""
    pass

class UpdateTransactionsRequest(BaseModel):
    processed_data: List[Dict[str, Any]]

class AssignClusterToLedgerRequest(BaseModel):
    """Payload for directly assigning a cluster's transactions to a ledger."""
    transactions: List[Dict[str, Any]]
    ledger_name: str
    add_to_known_ledgers: bool = False

class ClientUpdate(ClientBase):
    """Model used for updating an existing client's name."""
    pass

class RuleStat(BaseModel):
    """Defines the health statistics for a single rule."""
    total_matches: int
    last_used: Optional[str] = None

class RuleStatsResponse(BaseModel):
    """The response model for the rule stats endpoint."""
    stats: Dict[str, RuleStat]

class ClientModel(ClientBase):
    """A simple model for a single client's details."""
    id: str
    created_at: datetime
    updated_at: datetime
    
    class Config:
        from_attributes = True

class ClientModelWithCounts(ClientBase):
    """The full model representing a Client as returned from the API."""
    id: str
    created_at: datetime
    updated_at: datetime

    # Add the new count fields
    ledger_rule_count: int
    bank_statement_count: int
    bank_account_count: int
    known_ledger_count: int
    most_recent_statement_month: Optional[str] = None
    most_recent_statement_id: Optional[str] = None

    class Config:
        from_attributes = True

class StatementMetadata(BaseModel):
    """A summary model for a bank statement, used for listings."""
    id: str
    filename: str
    upload_date: datetime
    total_transactions: int
    matched_transactions: int
    bank_ledger_name: Optional[str] = None
    statement_period: Optional[str] = None
    status: str  # New field for "Completed" or "Needs Review"
    completion_percentage: float # New field for the percentage

    class Config:
        from_attributes = True

    
# Pydantic models for LedgerRule data
class LedgerRuleBase(BaseModel):
    """Defines the common attributes for a ledger rule."""
    client_id: str
    ledger_name: str
    regex_pattern: str
    sample_narrations: List[str] = []

class LedgerRule(LedgerRuleBase):
    """The model used when RETURNING a ledger rule from the API."""
    id: str
    created_at: datetime
    updated_at: datetime

    class Config:
        # --- THIS IS THE FIX ---
        from_attributes = True

class LedgerRuleUpdate(BaseModel):
    """Model for updating an existing ledger rule."""
    ledger_name: str
    regex_pattern: str

class BankStatement(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_id: str
    filename: str
    upload_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    column_mapping: Dict[str, Optional[str]]
    statement_format: str  # "single_amount_crdr" or "separate_credit_debit"
    raw_data: List[Dict]
    processed_data: Optional[List[Dict]] = None

class BankStatement(BaseModel):
    """
    Pydantic model for returning a BankStatement object from the API.
    """
    id: str
    client_id: str
    filename: str
    upload_date: datetime
    bank_account_id: Optional[str] = None # <-- ADD THIS LINE
    column_mapping: Dict[str, Optional[str]]
    statement_format: str
    
    # We include the raw_data and processed_data fields
    raw_data: List[Dict]
    processed_data: Optional[List[Dict]] = None

    class Config:
        from_attributes = True # Ensures it's compatible with our SQLAlchemy model

class ColumnMapping(BaseModel):
    date_column: str
    narration_column: str
    amount_column: Optional[str] = None
    credit_column: Optional[str] = None
    debit_column: Optional[str] = None
    balance_column: Optional[str] = None
    crdr_column: Optional[str] = None
    statement_format: str
    filtered_columns: Optional[List[str]] = []  # Columns to filter out from raw_data

class FileUploadResponse(BaseModel):
    file_id: str
    filename: str
    headers: List[str]
    preview_data: List[Dict]
    last_rows_data: List[Dict] = []  # Last 5 rows for preview
    suggested_mapping: Dict[str, str]
    problematic_rows: List[Dict] = []
    total_rows: int = 0
    problematic_row_indices: List[int] = []

class TransactionCluster(BaseModel):
    cluster_id: str
    narrations: List[str]
    suggested_regex: str
    keyword_patterns: List[str]
    confidence_score: float

class AIRegexRequest(BaseModel):
    narrations: List[str]
    existing_regex: Optional[str] = None
    use_local_llm: bool = False

class TallyVoucherData(BaseModel):
    client_id: str
    bank_ledger_name: str
    transactions: List[Dict]
    excluded_ledgers: List[str] = []

class KnownLedgerSummaryWithRuleCount(BaseModel):
    """A summary of a known ledger, including its rule count and last activity."""
    id: str
    ledger_name: str
    sample_count: int
    rule_count: int
    is_active: bool = True
    last_transaction_date: Optional[str] = None

class PaginatedLedgersResponse(BaseModel):
    """Response model for paginated known ledgers."""
    total_ledgers: int
    total_pages: int
    ledgers: List[KnownLedgerSummaryWithRuleCount]

# Pydantic models for BankAccount data
class BankAccountBase(BaseModel):
    """Base model for bank account, used for creation and updates."""
    client_id: str
    bank_name: str
    ledger_name: str
    contra_list: Optional[List[str]] = []
    filter_list: Optional[List[str]] = []

class BankAccountCreate(BankAccountBase):
    """Model used when creating a new bank account."""
    pass

class BankAccountUpdate(BankAccountBase):
    """Model used when updating an existing bank account."""
    pass

class BankAccountModel(BankAccountBase):
    """Full model representing a BankAccount as returned from the API."""
    id: str
    created_at: datetime
    updated_at: datetime
    class Config:
        # --- THIS IS THE FIX ---
        from_attributes = True

class DashboardStats(BaseModel):
    """Defines the metrics for the main dashboard."""
    total_clients: int
    statements_processed: int
    regex_patterns: int
    success_rate: float

class VoucherGenerationRequest(BaseModel):
    """Defines the user's selection for voucher export."""
    include_receipts: bool = True
    include_payments: bool = True
    include_contras: bool = True
    filename: str

class KnownLedgerSummary(BaseModel):
    """A summary of a known ledger for listing purposes."""
    id: str
    ledger_name: str
    sample_count: int
    is_active: bool = True


def compute_last_transaction_date(samples: Optional[List[Dict[str, Any]]]) -> Optional[str]:
    """
    Returns the most recent transaction date (ISO yyyy-mm-dd) from a list of samples.
    Samples may contain a 'date' field in various formats; we parse with pandas for robustness.
    """
    if not samples or not isinstance(samples, list):
        return None
    parsed_dates = []
    for sample in samples:
        date_str = sample.get("date") if isinstance(sample, dict) else None
        if not date_str:
            continue
        parsed = pd.to_datetime(date_str, dayfirst=True, errors="coerce")
        if pd.notna(parsed):
            parsed_dates.append(parsed)
    if not parsed_dates:
        return None
    return max(parsed_dates).date().isoformat()

class LearnLedgersRequest(BaseModel):
    """Request body for the learn-ledgers endpoint."""
    statement_ids: List[str]
    delay_ms: int = 100  # Delay between transactions for visualization

class AcceptLearnedLedgersRequest(BaseModel):
    """Request body for accepting/rejecting learned ledgers."""
    accepted_ledgers: List[str] = []  # Ledger names to accept
    rejected_ledgers: List[str] = []  # Ledger names to reject
    learning_results: Dict[str, Any] = {}  # The full results from learning phase

class ToggleLedgerActiveRequest(BaseModel):
    """Request body for toggling ledger active status."""
    is_active: bool

class SampleModel(BaseModel):
    """Defines the structure of a single transaction sample."""
    narration: str
    amount: float
    type: str # "Credit" or "Debit"
    date: Optional[str] = None  # DD/MM/YYYY format

class PaginatedSamplesResponse(BaseModel):
    """Response model for paginated samples."""
    total_samples: int
    samples: List[SampleModel]

class DeleteSampleRequest(BaseModel):
    """Request body for deleting a sample from a ledger."""
    index: Optional[int] = None
    sample: Optional[Dict[str, Any]] = None

class ReclassifySubsetRequest(BaseModel):
    """Defines the request body for re-classifying a subset of transactions."""
    transactions: List[Dict[str, Any]]

# Utility Functions
def detect_date_columns(headers: List[str]) -> List[str]:
    """Detect potential date columns"""
    # --- REPLACEMENT for the return statement ---
    normalized_headers = [h.lower().replace(" ", "").replace("_", "") for h in headers]
    date_keywords = ['date', 'dt', 'txndate', 'transactiondate', 'valuedate'] # Also normalized

    matched_headers = []
    for i, normalized_header in enumerate(normalized_headers):
        if any(keyword in normalized_header for keyword in date_keywords):
            matched_headers.append(headers[i]) # Append the ORIGINAL header
        
    return matched_headers

def detect_narration_columns(headers: List[str]) -> List[str]:
    """Detect potential narration/description columns"""
    narration_keywords = ['narration', 'description', 'desc', 'particulars', 'details', 'reference']
    return [h for h in headers if any(keyword in h.lower() for keyword in narration_keywords)]

def detect_amount_columns(headers: List[str]) -> List[str]:
    """Detect potential amount columns"""
    # Priority keywords - columns containing these will be prioritized
    priority_keywords = [
        'transaction amount',
        'transactionamount',
        'txn amount',
        'txnamount',
        'amount',
        'amt'
    ]
    
    # Secondary keywords - checked after priority keywords
    secondary_keywords = ['value', 'debit', 'credit', 'dr', 'cr', 'balance', 'bal']
    
    priority_matches = []
    secondary_matches = []
    date_matches = []  # Columns containing 'date' - lowest priority
    
    for h in headers:
        h_lower = h.lower()
        
        # Skip columns that contain 'date' - they should be deprioritized
        if 'date' in h_lower:
            # Still check if they match amount keywords, but put them in lowest priority
            if any(keyword in h_lower for keyword in priority_keywords + secondary_keywords):
                date_matches.append(h)
            continue
        
        # Check priority keywords first
        for keyword in priority_keywords:
            if keyword in h_lower:
                priority_matches.append(h)
                break
        else:
            # Only check secondary keywords if no priority match found
            if any(keyword in h_lower for keyword in secondary_keywords):
                secondary_matches.append(h)
    
    # Return priority matches first, then secondary matches, then date matches last
    return priority_matches + secondary_matches + date_matches

def suggest_column_mapping(headers: List[str]) -> Dict[str, str]:
    """Suggest column mapping based on headers"""
    mapping = {}
    
    date_cols = detect_date_columns(headers)
    if date_cols:
        mapping['date_column'] = date_cols[0]
    
    narration_cols = detect_narration_columns(headers)
    if narration_cols:
        mapping['narration_column'] = narration_cols[0]
    
    amount_cols = detect_amount_columns(headers)
    
    # Check for separate credit/debit columns
    credit_cols = [h for h in headers if 'credit' in h.lower() or h.lower() == 'cr']
    debit_cols = [h for h in headers if 'debit' in h.lower() or h.lower() == 'dr']
    
    if credit_cols and debit_cols:
        mapping['statement_format'] = 'separate_credit_debit'
        mapping['credit_column'] = credit_cols[0]
        mapping['debit_column'] = debit_cols[0]
    else:
        mapping['statement_format'] = 'single_amount_crdr'
        if amount_cols:
            mapping['amount_column'] = amount_cols[0]
        
        # Look for CR/DR indicator column
        crdr_cols = [h for h in headers if 'cr' in h.lower() and 'dr' in h.lower()]
        if crdr_cols:
            mapping['crdr_column'] = crdr_cols[0]
    
    # Balance column
    balance_cols = [h for h in headers if 'balance' in h.lower() or 'bal' in h.lower()]
    if balance_cols:
        mapping['balance_column'] = balance_cols[0]
    
    return mapping

def extract_keywords(text: str, devalued_keywords: set) -> set:
    """Extracts meaningful keywords from a narration string."""
    # This logic is ported from your regex_generator_v_4_clear_input.html
    tokens = re.split(r'[^A-Z0-9@]+', text.upper())
    meaningful_tokens = set()
    for token in tokens:
        if (
            token and
            len(token) > 2 and
            not re.match(r'\d{4,}', token) and
            not token.isdigit() and
            "**" not in token and
            token not in devalued_keywords
        ):
            meaningful_tokens.add(token)
    return meaningful_tokens

# In server.py

# --- FIND AND REPLACE THE ENTIRE get_dashboard_stats FUNCTION ---
@api_router.get("/dashboard-stats", response_model=DashboardStats)
async def get_dashboard_stats(db: AsyncSession = Depends(database.get_db)):
    """Calculates and returns key statistics for the main dashboard."""
    
    # --- THE FIX: Run queries sequentially, not in parallel ---
    total_clients = await db.scalar(select(func.count(models.Client.id)))
    statements_processed = await db.scalar(select(func.count(models.BankStatement.id)))
    regex_patterns = await db.scalar(select(func.count(models.LedgerRule.id)))
    # --- END OF FIX ---

    # Calculate success rate (this part of the logic is fine)
    total_transactions = 0
    matched_transactions = 0
    
    stmt_query = select(models.BankStatement.processed_data)
    result = await db.execute(stmt_query)
    all_processed_data = result.scalars().all()

    for processed_data in all_processed_data:
        if isinstance(processed_data, list):
            total_transactions += len(processed_data)
            matched_transactions += sum(1 for t in processed_data if t.get("matched_ledger") != "Suspense")
            
    success_rate = (matched_transactions / total_transactions * 100) if total_transactions > 0 else 0.0

    return DashboardStats(
        total_clients=total_clients or 0, # Add 'or 0' for safety
        statements_processed=statements_processed or 0,
        regex_patterns=regex_patterns or 0,
        success_rate=round(success_rate, 2)
    )
# --- END OF REPLACEMENT ---
# --- ADD REGEX VALIDATION HELPER FUNCTION ---
def validate_regex_pattern(pattern: str) -> Tuple[bool, str]:
    """
    Validates a regex pattern to ensure it's compatible with Python's re module.
    Returns (is_valid, error_message).
    """
    if not pattern or not pattern.strip():
        return False, "Regex pattern cannot be empty"
    
    try:
        # Try to compile the pattern to catch syntax errors
        re.compile(pattern)
        # Test with a simple search on empty string to catch runtime issues
        re.search(pattern, "")
        return True, ""
    except re.error as e:
        return False, f"Invalid regex pattern: {str(e)}"
    except Exception as e:
        return False, f"Unexpected error validating regex: {str(e)}"
# --- END OF VALIDATION HELPER ---

# --- REPLACEMENT for generate_regex_from_narrations ---
async def generate_regex_from_narrations(narrations: List[str], db: AsyncSession) -> str:
    """
    Generate a precise, order-aware regex pattern from narrations using common, meaningful keywords.
    """
    if not narrations:
        return ""

    # Fetch devalued keywords for filtering
    result = await db.execute(select(models.DevaluedKeyword.keyword))
    devalued_keywords_set = set(result.scalars().all())

    # 1. Find keywords that are common to ALL narrations in the cluster
    keyword_sets = [extract_keywords(n, devalued_keywords_set) for n in narrations]
    if not any(keyword_sets):
        return ".*"  # Fallback if no meaningful keywords are found at all
    common_keywords = set.intersection(*keyword_sets)

    # 2. If no keywords are common across all items, the cluster is weak.
    # Fallback to a simple regex based on the first item's first keyword.
    if not common_keywords:
        first_narration_keywords = sorted(list(extract_keywords(narrations[0], devalued_keywords_set)))
        if first_narration_keywords:
            return f".*\\b{re.escape(first_narration_keywords[0])}\\b.*"
        return ".*"

    # 3. Get an ordered list of all tokens from the first narration to use as a template.
    first_narration_ordered_tokens = re.split(r'[^A-Z0-9@]+', narrations[0].upper())

    # 4. Build the final list by picking common keywords in the order they appear in the template.
    ordered_common_keywords = []
    seen = set()
    for token in first_narration_ordered_tokens:
        if token in common_keywords and token not in seen:
            ordered_common_keywords.append(token)
            seen.add(token) # Prevents adding a keyword more than once

    # 5. Construct the final regex pattern from the ordered list.
    if ordered_common_keywords:
        pattern = ".*".join(f"\\b{re.escape(k)}\\b" for k in ordered_common_keywords)
        return f".*{pattern}.*" # Wrap with wildcards for flexibility
    
    # This is a final fallback, which should now be rarely needed.
    return f".*\\b{re.escape(sorted(list(common_keywords))[0])}\\b.*"

# In server.py

# --- ADD THIS ENTIRE NEW HELPER FUNCTION ---
# --- DEFINITIVE REPLACEMENT for pre_cluster_by_shared_keywords ---
async def pre_cluster_by_shared_keywords(
    transactions: List[Dict[str, Any]], narration_column: str, db: AsyncSession
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]: # <-- FIX 1: Correct type hint syntax
    """
    Finds high-confidence clusters based on shared keywords, ensuring no duplicates.
    Returns a tuple of (found_clusters, remaining_transactions).
    """
    result = await db.execute(select(models.DevaluedKeyword.keyword))
    devalued_keywords_set = set(result.scalars().all())
    
    keyword_map = defaultdict(list)
    for transaction in transactions:
        narration = str(transaction.get(narration_column, ""))
        keywords = extract_keywords(narration, devalued_keywords_set)
        for keyword in keywords:
            keyword_map[keyword].append(transaction)
            
    potential_clusters = []
    for keyword, mapped_transactions in keyword_map.items():
        if len(mapped_transactions) >= 2:
            potential_clusters.append(mapped_transactions)
            
    # Prioritize larger potential clusters first
    sorted_clusters = sorted(potential_clusters, key=len, reverse=True)
    
    final_clusters = []
    # --- FIX 2: Track using in-memory object ID, not the 'Srl' key ---
    processed_ids = set() 
    
    for cluster_group in sorted_clusters:
        # Filter out any transactions that have already been claimed by a larger cluster
        unclaimed_transactions = [
            t for t in cluster_group if id(t) not in processed_ids
        ]
        
        if len(unclaimed_transactions) >= 2:
            narrations = [str(t.get(narration_column, "")) for t in unclaimed_transactions]
            suggested_regex = await generate_regex_from_narrations(narrations, db)
            
            final_clusters.append({
                "cluster_id": str(uuid.uuid4()),
                "transactions": unclaimed_transactions,
                "suggested_regex": suggested_regex,
            })
            
            # Add the memory IDs of these transactions to the set so they can't be reused
            for t in unclaimed_transactions:
                processed_ids.add(id(t))

    # The remaining transactions are those whose memory IDs were never processed
    remaining_transactions = [
        t for t in transactions if id(t) not in processed_ids
    ]
    
    return final_clusters, remaining_transactions
# --- END OF REPLACEMENT ---
# --- FINAL REPLACEMENT for cluster_narrations ---
async def cluster_narrations(
    transactions: List[Dict[str, Any]], narration_column: str, db: AsyncSession, n_clusters: int = None
) -> List[Dict[str, Any]]:
    """Cluster similar transactions using TF-IDF and K-means, returning full transaction objects."""
    
    narrations = [str(t.get(narration_column, "")) for t in transactions]
    if not any(narrations):
        return []

    # Fetch devalued keywords for the vectorizer
    result = await db.execute(select(models.DevaluedKeyword.keyword))
    devalued_keywords = result.scalars().all()
    stop_words = [kw.lower() for kw in devalued_keywords] + ['english'] # FIX: Ensure all stop words are lowercase
    vectorizer = TfidfVectorizer(
        stop_words=stop_words,
        max_features=100,
        ngram_range=(1, 2),
        lowercase=True
    )

    try:
        tfidf_matrix = vectorizer.fit_transform(narrations)
        
        # Determine the optimal number of clusters if not specified
        if n_clusters is None:
            # Adjust clustering for smaller sample sizes
            num_samples = len(transactions)
        if n_clusters is None:
            if num_samples <= 5: # If 5 or fewer items, make each its own small cluster or group them tightly
                n_clusters = num_samples 
            else: # For larger groups, use the existing logic
                n_clusters = min(max(2, num_samples // 5), 10)
        # Ensure n_clusters is not zero if there are samples
        if num_samples > 0 and n_clusters == 0:
            n_clusters = 1
        
        kmeans = KMeans(n_clusters=n_clusters, random_state=42, n_init='auto')
        cluster_labels = kmeans.fit_predict(tfidf_matrix)
        
        # Group full transaction objects by their assigned cluster label
        clusters = defaultdict(list)
        for idx, label in enumerate(cluster_labels):
            clusters[label].append(transactions[idx])
        
        # Prepare the final list of cluster objects for the frontend
        cluster_objects = []
        for cluster_id, clustered_transactions in clusters.items():
            cluster_narrations_list = [str(t.get(narration_column, "")) for t in clustered_transactions]
            suggested_regex = await generate_regex_from_narrations(cluster_narrations_list, db)
            
            cluster_objects.append({
                "cluster_id": str(uuid.uuid4()),
                "transactions": clustered_transactions, # Return full objects
                "suggested_regex": suggested_regex,
            })
        return cluster_objects
    
    except Exception as e:
        logging.error(f"Clustering failed: {e}")
        # Fallback: Create a separate cluster for each transaction
        cluster_objects = []
        for transaction in transactions:
            narration = [str(transaction.get(narration_column, ""))]
            suggested_regex = await generate_regex_from_narrations(narration, db)
            cluster_objects.append({
                "cluster_id": str(uuid.uuid4()),
                "transactions": [transaction],
                "suggested_regex": suggested_regex
            })
        return cluster_objects


async def get_ai_improved_regex(narrations: List[str], existing_regex: str = None, use_local_llm: bool = False) -> str:
    """Get AI-improved regex pattern"""
    prompt = f"""
    You are a regex expert. Given these bank transaction narrations, create a precise regex pattern that matches all of them:
    
    Narrations:
    {chr(10).join(f"- {n}" for n in narrations[:10])}
    
    {"Current regex: " + existing_regex if existing_regex else ""}
    
    Requirements:
    1. Match ALL the provided narrations
    2. Be specific enough to avoid false positives
    3. Use word boundaries and case-insensitive patterns
    4. Focus on key identifying words, not variable parts like amounts or dates
    5. Return only the regex pattern, no explanation
    
    Regex pattern:
    """
    
    try:
        if use_local_llm and LOCAL_LLM_ENDPOINT:
            # Use local LLM
            response = requests.post(
                f"{LOCAL_LLM_ENDPOINT}/completions",
                json={
                    "prompt": prompt,
                    "max_tokens": 200,
                    "temperature": 0.1,
                    "stop": ["\n\n"]
                },
                timeout=30
            )
            if response.status_code == 200:
                return response.json()["choices"][0]["text"].strip()
        else:
            # Use OpenAI
            response = openai.chat.completions.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": "You are a regex expert helping with bank transaction pattern matching."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=200,
                temperature=0.1
            )
            return response.choices[0].message.content.strip()
    
    except Exception as e:
        logging.error(f"AI regex generation failed: {e}")
        # Fallback to basic regex generation
        return generate_regex_from_narrations(narrations)
    
    return generate_regex_from_narrations(narrations)

# API Routes

@api_router.get("/")
async def root():
    return {"message": "Tally Statement Processor API", "version": "1.0.0"}

# Client Management
@api_router.post("/clients", response_model=ClientModel)
async def create_client(client_data: ClientCreate, db: AsyncSession = Depends(database.get_db)):
    """Create a new client"""
    # Create a new SQLAlchemy model instance from the incoming data
    db_client = models.Client(**client_data.dict())
    
    # Add the new client to the session
    db.add(db_client)
    # Commit the transaction to save it to the database
    await db.commit()
    # Refresh the instance to get the default values from the DB (like created_at)
    await db.refresh(db_client)
    
    return db_client
# --- END OF REPLACEMENT ---

# --- CLEANER REPLACEMENT for get_clients ---
@api_router.get("/clients", response_model=List[ClientModelWithCounts])
async def get_clients(db: AsyncSession = Depends(database.get_db)):
    """Get all clients with their associated counts"""
    # (The complex query logic remains the same)
    ledger_rule_subquery = (
        select(models.LedgerRule.client_id, func.count(models.LedgerRule.id).label("ledger_rule_count"))
        .group_by(models.LedgerRule.client_id)
        .subquery()
    )
    # ... (other subqueries are the same) ...
    bank_statement_subquery = (
        select(models.BankStatement.client_id, func.count(models.BankStatement.id).label("bank_statement_count"))
        .group_by(models.BankStatement.client_id)
        .subquery()
    )
    bank_account_subquery = (
        select(models.BankAccount.client_id, func.count(models.BankAccount.id).label("bank_account_count"))
        .group_by(models.BankAccount.client_id)
        .subquery()
    )
    known_ledger_subquery = (
        select(models.KnownLedger.client_id, func.count(models.KnownLedger.id).label("known_ledger_count"))
        .group_by(models.KnownLedger.client_id)
        .subquery()
    )

    query = (
        select(
            models.Client,
            func.coalesce(ledger_rule_subquery.c.ledger_rule_count, 0).label("ledger_rule_count"),
            func.coalesce(bank_statement_subquery.c.bank_statement_count, 0).label("bank_statement_count"),
            func.coalesce(bank_account_subquery.c.bank_account_count, 0).label("bank_account_count"),
            func.coalesce(known_ledger_subquery.c.known_ledger_count, 0).label("known_ledger_count"),
        )
        .outerjoin(ledger_rule_subquery, models.Client.id == ledger_rule_subquery.c.client_id)
        .outerjoin(bank_statement_subquery, models.Client.id == bank_statement_subquery.c.client_id)
        .outerjoin(bank_account_subquery, models.Client.id == bank_account_subquery.c.client_id)
        .outerjoin(known_ledger_subquery, models.Client.id == known_ledger_subquery.c.client_id)
        .order_by(models.Client.name)
    )
    
    result = await db.execute(query)
    
    # Fetch all processed statements to determine most recent per client
    # This requires JSON parsing, so we do it separately
    processed_statements_query = (
        select(models.BankStatement)
        .where(models.BankStatement.processed_data.isnot(None))
        .order_by(models.BankStatement.upload_date.desc())
    )
    processed_statements_result = await db.execute(processed_statements_query)
    processed_statements = processed_statements_result.scalars().all()
    
    # Build a map of client_id -> most recent statement info
    client_most_recent = {}
    seen_clients = set()
    for stmt in processed_statements:
        if stmt.client_id not in seen_clients:
            seen_clients.add(stmt.client_id)
            # Extract month from processed_data
            month_str = None
            if isinstance(stmt.processed_data, list) and len(stmt.processed_data) > 0:
                # Get dates from processed_data (standardized format: DD/MM/YYYY)
                dates = []
                for t in stmt.processed_data:
                    date_str = t.get("Date")
                    if date_str:
                        try:
                            # Handle date string that might have extra text (split on space)
                            date_part = str(date_str).split(' ')[0] if ' ' in str(date_str) else str(date_str)
                            # Parse DD/MM/YYYY format using pandas
                            parsed_date = pd.to_datetime(date_part, format='%d/%m/%Y', errors='coerce')
                            if pd.notna(parsed_date):
                                dates.append(parsed_date)
                        except:
                            pass
                
                if dates:
                    # Use the maximum date to get the most recent month
                    max_date = max(dates)
                    month_str = max_date.strftime('%B, %Y')  # Format: "September, 2025"
            
            client_most_recent[stmt.client_id] = {
                "statement_id": stmt.id,
                "month": month_str
            }
    
    # Process the results more elegantly
    clients_with_counts = []
    for row in result.mappings(): # .mappings() gives us dict-like rows
        client_data = row['Client'].__dict__
        client_id = client_data['id']
        
        # Get most recent statement info for this client
        most_recent_info = client_most_recent.get(client_id, {})
        
        client_data.update({
            "ledger_rule_count": row['ledger_rule_count'],
            "bank_statement_count": row['bank_statement_count'],
            "bank_account_count": row['bank_account_count'],
            "known_ledger_count": row['known_ledger_count'],
            "most_recent_statement_month": most_recent_info.get("month"),
            "most_recent_statement_id": most_recent_info.get("statement_id"),
        })
        clients_with_counts.append(ClientModelWithCounts.model_validate(client_data))

    return clients_with_counts
# --- END OF CLEANER REPLACEMENT ---

# --- REPLACEMENT for get_client ---
@api_router.get("/clients/{client_id}", response_model=ClientModel)
async def get_client(client_id: str, db: AsyncSession = Depends(database.get_db)):
    """Get client by ID"""
    # Get a single client by its primary key (ID)
    client = await db.get(models.Client, client_id)
    
    # If the client doesn't exist, db.get() returns None
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
        
    return client
# --- END OF REPLACEMENT ---
# --- ADD THIS MISSING ENDPOINT ---
@api_router.put("/clients/{client_id}", response_model=ClientModel)
async def update_client(client_id: str, client_data: ClientUpdate, db: AsyncSession = Depends(database.get_db)):
    """Update an existing client's name"""
    db_client = await db.get(models.Client, client_id)
    if not db_client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Update the client object with the new data
    update_data = client_data.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_client, key, value)
        
    await db.commit()
    await db.refresh(db_client)
    return db_client
# --- END OF MISSING ENDPOINT ---

@api_router.delete("/clients/{client_id}", status_code=204)
async def delete_client(client_id: str, db: AsyncSession = Depends(database.get_db)):
    """Delete a client and all associated data"""
    db_client = await db.get(models.Client, client_id)
    if not db_client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    try:
        # Delete related records that don't have cascade delete configured
        # 1. Delete BankStatement records
        statements_query = select(models.BankStatement).where(
            models.BankStatement.client_id == client_id
        )
        statements_result = await db.execute(statements_query)
        statements = statements_result.scalars().all()
        for statement in statements:
            await db.delete(statement)
        
        # 2. Delete KnownLedger records
        known_ledgers_query = select(models.KnownLedger).where(
            models.KnownLedger.client_id == client_id
        )
        known_ledgers_result = await db.execute(known_ledgers_query)
        known_ledgers = known_ledgers_result.scalars().all()
        for ledger in known_ledgers:
            await db.delete(ledger)
        
        # 3. Delete ClassificationFeedback records
        feedback_query = select(models.ClassificationFeedback).where(
            models.ClassificationFeedback.client_id == client_id
        )
        feedback_result = await db.execute(feedback_query)
        feedback_records = feedback_result.scalars().all()
        for feedback in feedback_records:
            await db.delete(feedback)
        
        # 4. Delete the client (this will cascade delete BankAccount and LedgerRule records)
        await db.delete(db_client)
        await db.commit()
        return None  # Return None for 204 No Content response
        
    except Exception as e:
        await db.rollback()
        logging.error(f"Error deleting client {client_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error deleting client: {str(e)}")
# File Upload and Processing
# --- REPLACEMENT for upload_statement ---
@api_router.post("/upload-statement", response_model=FileUploadResponse)
async def upload_statement(file: UploadFile = File(...), db: AsyncSession = Depends(database.get_db)):
    """Upload and analyze bank statement"""
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Only Excel and CSV files are supported")
    
    try:
        content = await file.read()
        # --- START OF REPLACEMENT ---
        if file.filename.endswith('.csv'):
            # For CSV files, use read_csv with the index_col parameter
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))

        
        # --- END OF REPLACEMENT ---
        df = df.loc[:, ~df.columns.str.startswith('Unnamed:')]
        # Convert all data types in the DataFrame to strings to ensure JSON compatibility.

        df = df.dropna(how='all').reset_index(drop=True)

        # 1. Replace pandas' NaN/NaT with None, which is JSON-serializable as 'null'.
        # This is more robust than fillna('') as it preserves the "empty" nature.
        df = df.replace({np.nan: None, pd.NaT: None})

        # 2. Convert the entire clean DataFrame into a list of standard Python objects.
        # This is the key step that correctly converts numpy.int64 -> int, etc.
        sanitized_records = df.to_dict('records')
        
        # #region agent log
        with open(r"d:\Custom Tools\BSClassifier\.cursor\debug.log", "a", encoding="utf-8") as log_file:
            log_file.write(json.dumps({"sessionId": "debug-session", "runId": "run1", "hypothesisId": "validation", "location": "server.py:967", "message": "after sanitizing records", "data": {"record_count": len(sanitized_records), "sample_record_keys": list(sanitized_records[0].keys())[:10] if sanitized_records else []}, "timestamp": int(datetime.now(timezone.utc).timestamp() * 1000)}) + "\n")
        # #endregion

        headers = df.columns.tolist()
        preview_data = sanitized_records[:10] # Use the sanitized records for the preview
        # Get last 5 rows for preview (if there are more than 10 rows total)
        last_rows_data = []
        if len(sanitized_records) > 10:
            last_rows_data = sanitized_records[-5:]
        suggested_mapping = suggest_column_mapping(headers)
        
        # Scan for problematic rows immediately using suggested mapping
        problematic_rows = []
        problematic_row_indices = []
        
        if suggested_mapping:
            # Create a mapping dict for validation
            mapping_dict = {
                "date_column": suggested_mapping.get("date_column"),
                "narration_column": suggested_mapping.get("narration_column"),
                "amount_column": suggested_mapping.get("amount_column"),
                "credit_column": suggested_mapping.get("credit_column"),
                "debit_column": suggested_mapping.get("debit_column"),
                "statement_format": suggested_mapping.get("statement_format", "single_amount_crdr")
            }
            
            # Validate all rows using suggested mapping
            is_valid, error_msg, invalid_row_idx, problematic_rows_list = validate_transaction_data(
                sanitized_records, mapping_dict, return_all_problems=True
            )
            
            if problematic_rows_list:
                problematic_rows = problematic_rows_list
                problematic_row_indices = [row["row_index"] for row in problematic_rows_list]
                logging.info(f"Found {len(problematic_rows)} problematic rows during upload scan")
        
        # Store the sanitized data temporarily in the new table
        temp_file = models.TempFile(
            filename=file.filename,
            headers=headers,
            raw_data=sanitized_records
        )
        
        db.add(temp_file)
        await db.commit()
        await db.refresh(temp_file)
        
        return FileUploadResponse(
            file_id=temp_file.id, # Use the ID from the database
            filename=file.filename,
            headers=headers,
            preview_data=preview_data,
            last_rows_data=last_rows_data,
            suggested_mapping=suggested_mapping,
            problematic_rows=problematic_rows,
            total_rows=len(sanitized_records),
            problematic_row_indices=problematic_row_indices
        )
    
    except Exception as e:
        logging.error(f"File upload error: {e}")
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")
# --- END OF REPLACEMENT ---
def _clean_amount_string(value: Any) -> str:
    """
    Cleans a string to make it safely convertible to a number by removing
    commas and other non-numeric characters (except the decimal point).
    """
    if value is None:
        return "0"
    s = str(value)
    # Remove commas, then keep only digits, the first decimal point, and a leading minus sign.
    s = s.replace(",", "")
    # A simple regex to strip anything that's not a digit or a decimal point.
    # This is a safe way to handle currency symbols, etc.
    s = re.sub(r"[^0-9.-]", "", s)
    return s if s else "0"

def validate_transaction_data(transactions: List[Dict], mapping: Dict, required_fields: List[str] = None, return_all_problems: bool = False) -> Tuple[bool, Optional[str], Optional[int], Optional[List[Dict]]]:
    """
    Validates transaction data to ensure all required fields are present and non-None.
    
    Args:
        transactions: List of transaction dictionaries
        mapping: Column mapping dictionary
        required_fields: List of standardized field names to check (e.g., ['Narration', 'Date', 'Amount'])
                        If None, checks based on mapping type
        return_all_problems: If True, returns list of all problematic rows instead of stopping at first
    
    Returns:
        Tuple of (is_valid, error_message, row_index, problematic_rows)
        - is_valid: True if all rows are valid
        - error_message: User-friendly error message if invalid, None if valid
        - row_index: Index of the first invalid row (0-based), None if valid
        - problematic_rows: List of dicts with 'row_index', 'row_number', 'missing_fields', 'transaction_data'
    """
    if not transactions:
        return True, None, None, []
    
    # Determine required fields based on mapping if not provided
    if required_fields is None:
        required_fields = []
        if mapping.get("date_column"):
            required_fields.append("Date")
        if mapping.get("narration_column"):
            required_fields.append("Narration")
        if mapping.get("amount_column"):
            required_fields.append("Amount")
        elif mapping.get("statement_format") == "separate_credit_debit":
            required_fields.append("Amount (INR)")
    
    problematic_rows = []
    
    # Check each transaction
    for idx, transaction in enumerate(transactions):
        missing_fields = []
        
        for field in required_fields:
            value = None
            # For standardized fields, check directly
            if field in ["Date", "Narration", "Amount", "Amount (INR)", "CR/DR"]:
                value = transaction.get(field)
                # If not found as standardized, try raw mapping
                if value is None:
                    if field == "Date":
                        mapped_key = mapping.get("date_column")
                    elif field == "Narration":
                        mapped_key = mapping.get("narration_column")
                    elif field == "Amount":
                        mapped_key = mapping.get("amount_column")
                    elif field == "Amount (INR)":
                        # This is only for normalized data
                        pass
                    else:
                        mapped_key = None
                    if mapped_key:
                        value = transaction.get(mapped_key)
            else:
                # For raw fields, check via mapping
                mapped_key = mapping.get(field.lower() + "_column") or mapping.get(field)
                value = transaction.get(mapped_key) if mapped_key else None
            
            if value is None or (isinstance(value, str) and value.strip() == ""):
                missing_fields.append(field)
        
        if missing_fields:
            row_num = idx + 1  # 1-based for user display
            problematic_rows.append({
                "row_index": idx,
                "row_number": row_num,
                "missing_fields": missing_fields,
                "transaction_data": {k: str(v)[:100] if v is not None else None for k, v in list(transaction.items())[:10]}  # Limit data size
            })
            
            if not return_all_problems:
                # Return first problematic row (backward compatible)
                fields_str = ", ".join(missing_fields)
                error_msg = (
                    f"Row {row_num} in the statement is incomplete. "
                    f"Missing or empty required fields: {fields_str}. "
                    f"Please check the statement file and ensure all rows have complete data. "
                    f"If the issue persists, please report this to the administrator."
                )
                return False, error_msg, idx, problematic_rows
    
    if problematic_rows:
        # If we found problems but return_all_problems was True, create summary message
        fields_str = ", ".join(required_fields)
        error_msg = (
            f"Found {len(problematic_rows)} row(s) with missing or empty required fields ({fields_str}). "
            f"Please review and choose to filter them out or fix the data."
        )
        return False, error_msg, problematic_rows[0]["row_index"], problematic_rows
    
    return True, None, None, []
# --- END OF ADDITION ---

def normalize_transaction_data(
    raw_data: List[Dict], mapping: "ColumnMapping"
) -> List[Dict]:
    """
    Normalizes raw transaction data from separate credit/debit columns
    into a single amount column and a CR/DR indicator column.
    """
    if mapping.statement_format != "separate_credit_debit":
        return raw_data  # No changes needed if not in the separate format

    normalized_data = []
    credit_col = mapping.credit_column
    debit_col = mapping.debit_column

    if not credit_col or not debit_col:
        # If columns are not defined, we cannot proceed with normalization.
        return raw_data

    for row in raw_data:
        new_row = row.copy()
        try:
            # Safely convert to numeric, coercing errors to NaN, then to 0.0
            credit_val = pd.to_numeric(_clean_amount_string(row.get(credit_col)), errors='coerce')
            debit_val = pd.to_numeric(_clean_amount_string(row.get(debit_col)), errors='coerce')
            credit_val = credit_val if pd.notna(credit_val) else 0.0
            debit_val = debit_val if pd.notna(debit_val) else 0.0
            
            # Logic to determine Amount and CR/DR
            if credit_val > 0:
                # Convert from numpy.float64 to a standard Python float
                new_row["Amount (INR)"] = float(credit_val)
                new_row["CR/DR"] = "CR"
            elif debit_val > 0:
                # Convert from numpy.float64 to a standard Python float
                new_row["Amount (INR)"] = float(debit_val)
                new_row["CR/DR"] = "DR"
            else:
                # Handle rows with 0 in both, or non-standard entries
                new_row["Amount (INR)"] = 0.0
                new_row["CR/DR"] = "N/A"

            # Remove original credit/debit columns to avoid duplication
            if credit_col in new_row:
                del new_row[credit_col]
            if debit_col in new_row:
                del new_row[debit_col]

            normalized_data.append(new_row)

        except (ValueError, TypeError):
            # If conversion fails for a row, append it as is but log it.
            logging.warning(f"Could not normalize row: {row}. Appending as is.")
            normalized_data.append(row)
            continue
            
    return normalized_data

@api_router.post("/statements/{statement_id}/reclassify-subset", response_model=List[Dict[str, Any]])
async def reclassify_subset(
    statement_id: str,
    request: ReclassifySubsetRequest,
    db: AsyncSession = Depends(database.get_db)
):
    """
    Re-applies all regex rules to a specific subset of transactions and returns the result
    without saving it to the database.
    """
    statement = await db.get(models.BankStatement, statement_id)
    if not statement:
        raise HTTPException(status_code=404, detail="Statement not found")

    # Fetch all rules for the client
    query = select(models.LedgerRule).where(models.LedgerRule.client_id == statement.client_id)
    result = await db.execute(query)
    patterns = result.scalars().all()

    # Create a list to hold the results
    reclassified_transactions = []

    # Iterate ONLY through the transactions provided in the request
    for transaction in request.transactions:
        # The frontend uses 'Narration' as the standardized key
        narration = str(transaction.get("Narration", ""))
        matched = False
        
        # Apply the same matching logic as the main classification function
        for pattern in patterns:
            if re.search(pattern.regex_pattern, narration, re.IGNORECASE):
                # Update the ledger and ensure user_confirmed is false
                transaction["matched_ledger"] = pattern.ledger_name
                transaction["user_confirmed"] = False 
                reclassified_transactions.append(transaction)
                matched = True
                break
        
        if not matched:
            # If no rule matches, it becomes a soft suspense item
            transaction["matched_ledger"] = "Suspense"
            transaction["user_confirmed"] = False
            reclassified_transactions.append(transaction)

    return reclassified_transactions

# --- REPLACEMENT for confirm_column_mapping ---
@api_router.post("/validate-mapping/{file_id}")
async def validate_column_mapping(file_id: str, mapping: ColumnMapping, db: AsyncSession = Depends(database.get_db)):
    """Validate column mapping and return problematic rows"""
    temp_file = await db.get(models.TempFile, file_id)
    if not temp_file:
        raise HTTPException(status_code=404, detail="File not found or has expired")

    try:
        mapping_dict = mapping.dict()
        
        # Validate raw data and get all problematic rows
        is_valid, error_msg, invalid_row_idx, problematic_rows = validate_transaction_data(
            temp_file.raw_data, mapping_dict, return_all_problems=True
        )
        
        if is_valid:
            return {
                "is_valid": True,
                "problematic_rows": [],
                "total_rows": len(temp_file.raw_data),
                "valid_rows": len(temp_file.raw_data)
            }
        
        # Return problematic rows information
        return {
            "is_valid": False,
            "problematic_rows": problematic_rows,
            "total_rows": len(temp_file.raw_data),
            "valid_rows": len(temp_file.raw_data) - len(problematic_rows),
            "error_message": error_msg
        }
    
    except Exception as e:
        logging.error(f"Validation error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error validating mapping: {str(e)}")

@api_router.post("/confirm-mapping/{file_id}")
async def confirm_column_mapping(
    file_id: str, 
    mapping: ColumnMapping, 
    client_id: str = Query(...), 
    bank_account_id: str = Query(...),
    filter_problematic_rows: bool = Query(False, description="Filter out problematic rows before processing"),
    db: AsyncSession = Depends(database.get_db)
):
    """Confirm column mapping and process statement"""
    temp_file = await db.get(models.TempFile, file_id)
    if not temp_file:
        raise HTTPException(status_code=404, detail="File not found or has expired")

    try:
        # --- START OF NEW LOGIC ---
        mapping_dict = mapping.dict()
        
        # Get problematic rows if filtering is requested
        problematic_row_indices = set()
        if filter_problematic_rows:
            is_valid, error_msg, invalid_row_idx, problematic_rows = validate_transaction_data(
                temp_file.raw_data, mapping_dict, return_all_problems=True
            )
            if problematic_rows:
                problematic_row_indices = {row["row_index"] for row in problematic_rows}
                logging.info(f"Filtering out {len(problematic_row_indices)} problematic rows")
        
        # Filter out problematic rows if requested
        raw_data_to_process = temp_file.raw_data
        if problematic_row_indices:
            raw_data_to_process = [
                row for idx, row in enumerate(temp_file.raw_data) 
                if idx not in problematic_row_indices
            ]
            logging.info(f"Processed {len(raw_data_to_process)} rows after filtering (removed {len(problematic_row_indices)} problematic rows)")
        
        # Filter out user-selected columns from raw_data
        filtered_columns = mapping.filtered_columns or []
        if filtered_columns:
            # Remove filtered columns from each row in raw_data
            raw_data_to_process = [
                {k: v for k, v in row.items() if k not in filtered_columns}
                for row in raw_data_to_process
            ]
            logging.info(f"Filtered out {len(filtered_columns)} columns: {', '.join(filtered_columns)}")
        
        # Normalize the data before creating the statement object
        processed_raw_data = normalize_transaction_data(raw_data_to_process, mapping)
        
        # Validate normalized data as well (should be valid if we filtered)
        final_mapping = mapping_dict.copy()
        if mapping.statement_format == "separate_credit_debit":
            final_mapping["amount_column"] = "Amount (INR)"
            final_mapping["crdr_column"] = "CR/DR"
            # Set original columns to None as they no longer exist
            final_mapping["credit_column"] = None
            final_mapping["debit_column"] = None
        
        # Only validate if we didn't filter (if we filtered, data should be clean)
        if not filter_problematic_rows:
            is_valid, error_msg, invalid_row_idx, _ = validate_transaction_data(processed_raw_data, final_mapping, return_all_problems=False)
            if not is_valid:
                await db.rollback()
                logging.error(f"Invalid normalized transaction data at row {invalid_row_idx}: {error_msg}")
                raise HTTPException(status_code=400, detail=error_msg)
        # --- END OF NEW LOGIC ---

        # Create a permanent BankStatement record
        statement = models.BankStatement(
            client_id=client_id,
            bank_account_id=bank_account_id,
            filename=temp_file.filename,
            column_mapping=final_mapping,  # Use the potentially updated mapping
            statement_format=mapping.statement_format,
            raw_data=processed_raw_data  # Use the normalized data
        )
        
        db.add(statement)
        await db.delete(temp_file)
        await db.commit()
        await db.refresh(statement)
        
        return {
            "message": "Statement processed successfully", 
            "statement_id": statement.id,
            "rows_filtered": len(problematic_row_indices) if filter_problematic_rows else 0,
            "rows_processed": len(processed_raw_data)
        }
    
    except HTTPException:
        # Re-raise HTTP exceptions (including our validation errors) as-is
        raise
    except Exception as e:
        await db.rollback()
        logging.error(f"Mapping confirmation error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error processing mapping: {str(e)}")

# Regex Pattern Management
# --- REPLACEMENT for create_regex_pattern ---
@api_router.post("/ledger-rules", response_model=LedgerRule)
async def create_ledger_rule(rule_data: LedgerRuleBase, db: AsyncSession = Depends(database.get_db)):
        """Create a new ledger rule"""
        # --- VALIDATE REGEX PATTERN BEFORE SAVING ---
        is_valid, error_message = validate_regex_pattern(rule_data.regex_pattern)
        if not is_valid:
            raise HTTPException(status_code=400, detail=error_message)
        # --- END OF VALIDATION ---
        
        db_rule = models.LedgerRule(**rule_data.dict())
        
        db.add(db_rule)
        await db.commit()
        await db.refresh(db_rule)
        # --- START OF ADDITION ---
        # Now, create and save the corresponding feedback entry.
        feedback_entry = models.ClassificationFeedback(
            client_id=db_rule.client_id,
            ledger_name=db_rule.ledger_name,
            source_narrations=db_rule.sample_narrations,
            regex_pattern=rule_data.regex_pattern # Store the regex directly
        )
        # 3. Add both objects to the session.
        db.add(db_rule)
        db.add(feedback_entry)
    
        try:
            # 4. Commit the session ONCE to save both objects atomically.
            await db.commit()
        except Exception as e:
            # If anything fails, roll back both changes.
            await db.rollback()
            raise HTTPException(status_code=500, detail=f"Database error: {e}")

        # 5. Refresh the rule object to load its generated ID and timestamps.
        await db.refresh(db_rule)
        
        # 6. Return the rule object. The session is still active, so no error will occur.
        return db_rule

# --- REPLACEMENT for get_client_regex_patterns ---
@api_router.get("/ledger-rules/{client_id}", response_model=List[LedgerRule])
async def get_client_ledger_rules(client_id: str, db: AsyncSession = Depends(database.get_db)):
        """Get all ledger rules for a client"""
        query = select(models.LedgerRule).where(models.LedgerRule.client_id == client_id)
        
        result = await db.execute(query)
        rules = result.scalars().all()
        
        return rules
# --- END OF REPLACEMENT ---
# --- ADD THIS ENTIRE HELPER FUNCTION ---
def create_standardized_transaction(transaction: Dict, mapping: Dict) -> Dict:
    """
    Creates a new, clean transaction dictionary containing only the essential
    columns under standardized key names.
    """
    standardized = {}
    # --- START: Date Normalization Logic ---
    raw_date_str = transaction.get(mapping.get("date_column"))
    if raw_date_str:
        # Use pandas for robust parsing, hinting that the day is first.
        # errors='coerce' will return NaT (Not a Time) for unparseable dates.
        parsed_date = pd.to_datetime(raw_date_str, dayfirst=True, errors='coerce')
        
        # Check if parsing was successful. pd.notna checks for NaT.
        if pd.notna(parsed_date):
            # Reformat to our single, consistent DD/MM/YYYY standard, stripping all time info.
            standardized["Date"] = parsed_date.strftime('%d/%m/%Y')
        else:
            # If parsing fails, use the original string as a safe fallback.
            standardized["Date"] = str(raw_date_str)
    else:
        standardized["Date"] = None # Handle cases where the date is missing.
    # --- END: Date Normalization Logic ---
    # Map essential columns using the provided mapping
    narration_col_key = mapping.get("narration_column")
    raw_narration = transaction.get(narration_col_key)
    # #region agent log
    with open(r"d:\Custom Tools\BSClassifier\.cursor\debug.log", "a", encoding="utf-8") as log_file:
        log_file.write(json.dumps({"sessionId": "debug-session", "runId": "run1", "hypothesisId": "B,C", "location": "server.py:1331", "message": "creating standardized Narration", "data": {"narration_col_key": narration_col_key, "raw_narration": raw_narration, "raw_narration_type": str(type(raw_narration)), "transaction_has_key": narration_col_key in transaction if narration_col_key else False, "transaction_keys_sample": list(transaction.keys())[:10]}, "timestamp": int(datetime.now(timezone.utc).timestamp() * 1000)}) + "\n")
    # #endregion
    # Ensure Narration is never None - use empty string as fallback
    standardized["Narration"] = str(raw_narration) if raw_narration is not None else ""
    standardized["Amount"] = transaction.get(mapping.get("amount_column"))
    raw_cr_dr = transaction.get(mapping.get("crdr_column"), "")
    # Clean and normalize it immediately
    clean_cr_dr = str(raw_cr_dr).strip().replace(".", "").upper()
    standardized["CR/DR"] = clean_cr_dr

    # Only include Balance if it was mapped by the user
    if mapping.get("balance_column"):
        standardized["Balance"] = transaction.get(mapping.get("balance_column"))
        
    return standardized
# --- END OF ADDITION ---
# --- ADD THESE TWO NEW ENDPOINTS ---
@api_router.put("/ledger-rules/{rule_id}", response_model=LedgerRule)
async def update_ledger_rule(rule_id: str, rule_data: LedgerRuleUpdate, db: AsyncSession = Depends(database.get_db)):
    """Update a ledger rule by its ID."""
    db_rule = await db.get(models.LedgerRule, rule_id)
    if not db_rule:
        raise HTTPException(status_code=404, detail="Ledger rule not found")
    
    # --- VALIDATE REGEX PATTERN BEFORE UPDATING ---
    is_valid, error_message = validate_regex_pattern(rule_data.regex_pattern)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_message)
    # --- END OF VALIDATION ---
    
    # Update the rule object with the new data
    update_data = rule_data.dict()
    for key, value in update_data.items():
        setattr(db_rule, key, value)
        
    await db.commit()
    await db.refresh(db_rule)
    return db_rule

@api_router.delete("/ledger-rules/{rule_id}", status_code=204)
async def delete_ledger_rule(rule_id: str, db: AsyncSession = Depends(database.get_db)):
    """Delete a ledger rule by its ID."""
    db_rule = await db.get(models.LedgerRule, rule_id)
    if not db_rule:
        raise HTTPException(status_code=404, detail="Ledger rule not found")
    
    await db.delete(db_rule)
    await db.commit()
    return None # Return None for 204 No Content response
# --- END OF ADDITION ---
# In server.py
@api_router.get("/clients/{client_id}/rule-stats", response_model=RuleStatsResponse)
async def get_rule_stats(client_id: str, db: AsyncSession = Depends(database.get_db)):
    """Calculates on-demand health statistics for all rules of a given client."""
    
    # Step 1: Fetch all statements for the client
    stmt_query = select(models.BankStatement).where(models.BankStatement.client_id == client_id)
    result = await db.execute(stmt_query)
    statements = result.scalars().all()

    # Step 2: Initialize a structure to hold our calculations
    # We use ledger_name as the key.
    stats = defaultdict(lambda: {"total_matches": 0, "last_used_dt": None})

    # Step 3: Iterate through all transactions of all statements
    for statement in statements:
        if not isinstance(statement.processed_data, list):
            continue
        
        for transaction in statement.processed_data:
            ledger_name = transaction.get("matched_ledger")
            
            # We only care about transactions matched to a specific rule, not Suspense
            if not ledger_name or ledger_name == "Suspense":
                continue

            # Increment the match counter
            stats[ledger_name]["total_matches"] += 1

            # Update the last used date
            date_str = transaction.get("Date")
            if date_str:
                try:
                    # Parse the date string into a datetime object for comparison
                    # This handles formats like 'DD/MM/YYYY HH:MM:SS'
                    transaction_dt = datetime.strptime(date_str.split(' ')[0], '%d/%m/%Y')
                    
                    if (stats[ledger_name]["last_used_dt"] is None or 
                        transaction_dt > stats[ledger_name]["last_used_dt"]):
                        stats[ledger_name]["last_used_dt"] = transaction_dt
                except (ValueError, TypeError):
                    continue # Ignore rows with unparseable dates

    # Step 4: Format the results for the final JSON response
    final_stats = {
        ledger: RuleStat(
            total_matches=data["total_matches"],
            last_used=data["last_used_dt"].strftime('%d %b %Y') if data["last_used_dt"] else None
        )
        for ledger, data in stats.items()
    }

    return RuleStatsResponse(stats=final_stats)
# --- END OF ADDITION ---

# --- FIND AND REPLACE THE ENTIRE classify-transactions FUNCTION ---
@api_router.post("/classify-transactions/{statement_id}")
async def classify_transactions(
    statement_id: str, 
    db: AsyncSession = Depends(database.get_db),
    force_reclassify: bool = Query(False, description="Force re-matching of rules against pending transactions")
):
    """
    Classify transactions. This endpoint is STATE-AWARE and now ONLY uses processed_data.
    """
    statement = await db.get(models.BankStatement, statement_id)
    if not statement:
        raise HTTPException(status_code=404, detail="Statement not found")

    mapping = statement.column_mapping
    narration_col = mapping.get("narration_column")
    
    classified_transactions = statement.processed_data if isinstance(statement.processed_data, list) else []

    if not classified_transactions: # First run
        # Validate raw_data before processing
        raw_data = statement.raw_data or []
        if raw_data:
            is_valid, error_msg, invalid_row_idx, _ = validate_transaction_data(raw_data, mapping, return_all_problems=False)
            if not is_valid:
                logging.error(f"Invalid raw_data during first classification: {error_msg}")
                raise HTTPException(
                    status_code=400,
                    detail=error_msg
                )
        
        query = select(models.LedgerRule).where(models.LedgerRule.client_id == statement.client_id)
        result = await db.execute(query)
        patterns = result.scalars().all()
        # #region agent log
        with open(r"d:\Custom Tools\BSClassifier\.cursor\debug.log", "a", encoding="utf-8") as log_file:
            log_file.write(json.dumps({"sessionId": "debug-session", "runId": "run1", "hypothesisId": "B,C", "location": "server.py:1475", "message": "first run classification started", "data": {"raw_data_count": len(raw_data), "narration_col": narration_col, "mapping": str(mapping)}, "timestamp": int(datetime.now(timezone.utc).timestamp() * 1000)}) + "\n")
        # #endregion
        for transaction in raw_data:
            standardized_transaction = create_standardized_transaction(transaction, mapping)
            # #region agent log
            with open(r"d:\Custom Tools\BSClassifier\.cursor\debug.log", "a", encoding="utf-8") as log_file:
                log_file.write(json.dumps({"sessionId": "debug-session", "runId": "run1", "hypothesisId": "B,C", "location": "server.py:1483", "message": "after create_standardized_transaction", "data": {"standardized_narration": standardized_transaction.get("Narration"), "standardized_narration_type": str(type(standardized_transaction.get("Narration"))), "raw_transaction_keys": list(transaction.keys())[:10], "narration_col_value": transaction.get(narration_col) if narration_col else None}, "timestamp": int(datetime.now(timezone.utc).timestamp() * 1000)}) + "\n")
            # #endregion
            
            # Ensure narration is never None - use empty string as fallback
            narration = standardized_transaction.get("Narration")
            if narration is None:
                narration = str(transaction.get(narration_col, "")) if narration_col else ""
            else:
                narration = str(narration) if narration else ""
            
            # Double-check: if narration is still empty/None after all attempts, skip this transaction
            if not narration or narration.strip() == "":
                logging.warning(f"Skipping transaction with empty narration: {transaction}")
                continue
            
            matched = False
            for pattern in patterns:
                try:
                    if re.search(pattern.regex_pattern, narration, re.IGNORECASE):
                        standardized_transaction["matched_ledger"] = pattern.ledger_name
                        classified_transactions.append(standardized_transaction)
                        matched = True
                        break
                except re.error as e:
                    # Log invalid patterns (shouldn't happen with validation, but keep as safety net)
                    logging.warning(f"Invalid regex pattern for ledger {pattern.ledger_name} (ID: {pattern.id}): {e}. Pattern: {pattern.regex_pattern[:100]}")
                    continue
            if not matched:
                standardized_transaction["matched_ledger"] = "Suspense"
                classified_transactions.append(standardized_transaction)
        statement.processed_data = classified_transactions
        await db.commit()
        await db.refresh(statement)

    if force_reclassify: # Intelligent Merge
        # Validate processed_data before reclassification
        if classified_transactions:
            is_valid, error_msg, invalid_row_idx, _ = validate_transaction_data(classified_transactions, mapping, return_all_problems=False)
            if not is_valid:
                logging.error(f"Invalid processed_data during reclassification: {error_msg}")
                raise HTTPException(
                    status_code=400,
                    detail=error_msg
                )
        
        items_to_recheck = [t for t in classified_transactions if t.get("matched_ledger") == "Suspense" and not t.get("user_confirmed")]
        # #region agent log
        with open(r"d:\Custom Tools\BSClassifier\.cursor\debug.log", "a", encoding="utf-8") as log_file:
            log_file.write(json.dumps({"sessionId": "debug-session", "runId": "run1", "hypothesisId": "A,B,C,D,E", "location": "server.py:1390", "message": "force_reclassify block entered", "data": {"force_reclassify": force_reclassify, "items_to_recheck_count": len(items_to_recheck), "mapping": str(mapping), "narration_col": narration_col}, "timestamp": int(datetime.now(timezone.utc).timestamp() * 1000)}) + "\n")
        # #endregion
        if items_to_recheck:
            query = select(models.LedgerRule).where(models.LedgerRule.client_id == statement.client_id)
            result = await db.execute(query)
            patterns = result.scalars().all()
            reclassification_updates = {}
            invalid_items = []
            for item_idx, item in enumerate(items_to_recheck):
                # #region agent log
                with open(r"d:\Custom Tools\BSClassifier\.cursor\debug.log", "a", encoding="utf-8") as log_file:
                    log_file.write(json.dumps({"sessionId": "debug-session", "runId": "run1", "hypothesisId": "A,B,C,D,E", "location": "server.py:1396", "message": "processing item in recheck loop", "data": {"item_keys": list(item.keys()), "narration_value": item.get('Narration'), "narration_type": str(type(item.get('Narration'))), "narration_is_none": item.get('Narration') is None, "item_sample": {k: str(v)[:50] if v is not None else None for k, v in list(item.items())[:5]}}, "timestamp": int(datetime.now(timezone.utc).timestamp() * 1000)}) + "\n")
                # #endregion
                narration_val = item.get('Narration')
                
                # Guard against None or empty narrations
                if narration_val is None or (isinstance(narration_val, str) and narration_val.strip() == ""):
                    invalid_items.append(item_idx)
                    # #region agent log
                    with open(r"d:\Custom Tools\BSClassifier\.cursor\debug.log", "a", encoding="utf-8") as log_file:
                        log_file.write(json.dumps({"sessionId": "debug-session", "runId": "run1", "hypothesisId": "A", "location": "server.py:1405", "message": "skipping item with None/empty narration", "data": {"item_idx": item_idx, "item_keys": list(item.keys())}, "timestamp": int(datetime.now(timezone.utc).timestamp() * 1000)}) + "\n")
                    # #endregion
                    continue
                
                for pattern in patterns:
                    try:
                        # #region agent log
                        with open(r"d:\Custom Tools\BSClassifier\.cursor\debug.log", "a", encoding="utf-8") as log_file:
                            log_file.write(json.dumps({"sessionId": "debug-session", "runId": "run1", "hypothesisId": "A", "location": "server.py:1410", "message": "before re.search call", "data": {"narration_value": narration_val, "narration_is_none": narration_val is None, "pattern_id": pattern.id, "pattern_ledger": pattern.ledger_name}, "timestamp": int(datetime.now(timezone.utc).timestamp() * 1000)}) + "\n")
                        # #endregion
                        if re.search(pattern.regex_pattern, narration_val, re.IGNORECASE):
                            reclassification_updates[narration_val] = pattern.ledger_name
                            break
                    except re.error as e:
                        logging.warning(f"Invalid regex pattern for ledger {pattern.ledger_name} (ID: {pattern.id}) during reclassification: {e}. Pattern: {pattern.regex_pattern[:100]}")
                        continue
                    except TypeError as e:
                        # #region agent log
                        with open(r"d:\Custom Tools\BSClassifier\.cursor\debug.log", "a", encoding="utf-8") as log_file:
                            log_file.write(json.dumps({"sessionId": "debug-session", "runId": "run1", "hypothesisId": "A", "location": "server.py:1545", "message": "TypeError in re.search", "data": {"error": str(e), "narration_value": item.get('Narration'), "narration_type": str(type(item.get('Narration')))}, "timestamp": int(datetime.now(timezone.utc).timestamp() * 1000)}) + "\n")
                        # #endregion
                        invalid_items.append(item_idx)
                        break
            
            # Check for invalid items after processing all items
            if invalid_items:
                invalid_count = len(invalid_items)
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Found {invalid_count} transaction(s) with missing or invalid narration data. "
                        f"This indicates malformed rows in the statement. "
                        f"Please check the statement file and ensure all rows have complete data. "
                        f"If the issue persists, please report this to the administrator."
                    )
                )
            if reclassification_updates:
                for i, t in enumerate(classified_transactions):
                    if t['Narration'] in reclassification_updates:
                        classified_transactions[i]['matched_ledger'] = reclassification_updates[t['Narration']]
                statement.processed_data = classified_transactions
                await db.commit()
                await db.refresh(statement)

    # --- THIS IS THE CRITICAL CHANGE ---
    # We now identify the transactions to cluster directly from our master list.
    transactions_to_cluster = [
        t for t in classified_transactions
        if t.get("matched_ledger") == "Suspense" and not t.get("user_confirmed")
    ]
    # --- END OF CHANGE ---

    # The clustering functions will now receive standardized objects.
    # We pass the STANDARDIZED narration key "Narration" to them.
    pre_clusters, remaining_for_ml = await pre_cluster_by_shared_keywords(transactions_to_cluster, "Narration", db)
    
    debit_transactions = [t for t in remaining_for_ml if str(t.get("CR/DR","")).strip().upper() == "DR"]
    credit_transactions = [t for t in remaining_for_ml if str(t.get("CR/DR","")).strip().upper() == "CR"]
    
    debit_clusters = await cluster_narrations(debit_transactions, "Narration", db)
    credit_clusters = await cluster_narrations(credit_transactions, "Narration", db)
    final_clusters = pre_clusters + debit_clusters + credit_clusters
    
    # The total transactions should be the count of classified_transactions,
    # not raw_data, since classified_transactions is what we actually work with
    # and what the UI displays
    total_transactions_count = len(classified_transactions)
    
    return {
        "classified_transactions": classified_transactions,
        "unmatched_clusters": final_clusters,
        "total_transactions": total_transactions_count,
        "matched_transactions": len([t for t in classified_transactions if t.get("matched_ledger") != "Suspense"]),
        "unmatched_transactions": len(transactions_to_cluster)
    }
# --- END OF REPLACEMENT ---
# In server.py, after the classify_transactions function

# --- ADD THIS ENTIRE ENDPOINT ---
@api_router.post("/statements/{statement_id}/update-transactions")
async def update_transactions(statement_id: str, request: UpdateTransactionsRequest, db: AsyncSession = Depends(database.get_db)):
    """Updates the processed_data for a given statement using a canonical-key merge to avoid duplicates."""
    statement = await db.get(models.BankStatement, statement_id)
    if not statement:
        raise HTTPException(status_code=404, detail="Statement not found")

    existing = statement.processed_data or []
    incoming = request.processed_data or []

    def _normalize_narration(n):
        if not n:
            return ""
        s = str(n).strip().lower()
        s = re.sub(r'\s+', ' ', s)
        return s

    def _normalize_date(d):
        if not d:
            return ""
        s = str(d).split(' ')[0].strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y"):
            try:
                return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
            except Exception:
                continue
        return s.lower()

    def _normalize_amount(a):
        if a is None:
            return 0.0
        try:
            s = str(a).replace(',', '').strip()
            return float(s)
        except Exception:
            try:
                return float(re.sub(r'[^\d.-]', '', str(a)))
            except Exception:
                return 0.0

    def canonical_key(tx):
        amount = tx.get("Amount") if "Amount" in tx else tx.get("Amount (INR)") if "Amount (INR)" in tx else tx.get("amount") if "amount" in tx else tx.get("amt")
        return f"{_normalize_narration(tx.get('Narration') or tx.get('narration'))}||{_normalize_date(tx.get('Date') or tx.get('date'))}||{_normalize_amount(amount)}"

    # Build incoming map keyed by canonical key (last one wins for duplicates in incoming)
    incoming_map = {}
    for tx in incoming:
        incoming_map[canonical_key(tx)] = tx

    updated_count = 0
    appended_count = 0
    processed_incoming_keys = set()

    # Merge: prefer updating existing entries in-place (preserve order)
    merged = []
    for orig in existing:
        key = canonical_key(orig)
        if key in incoming_map:
            incoming_tx = incoming_map[key]
            # Merge: incoming fields override existing ones (but keep other fields)
            merged_tx = {**orig, **incoming_tx}

            # Safety rule: if transaction is (or becomes) Suspense, ensure user_confirmed is explicitly FALSE
            # unless the incoming payload explicitly set user_confirmed (in which case honor it).
            if str(merged_tx.get('matched_ledger', '')).strip().lower() == 'suspense':
                if 'user_confirmed' in incoming_tx:
                    merged_tx['user_confirmed'] = bool(incoming_tx['user_confirmed'])
                else:
                    merged_tx['user_confirmed'] = False

            # If incoming explicitly set user_confirmed (true/false), ensure it is honored regardless of ledger
            elif 'user_confirmed' in incoming_tx:
                merged_tx['user_confirmed'] = bool(incoming_tx['user_confirmed'])

            merged.append(merged_tx)
            processed_incoming_keys.add(key)
            updated_count += 1
        else:
            merged.append(orig)

    # Append any incoming transactions that did not match existing keys
    for key, tx in incoming_map.items():
        if key in processed_incoming_keys:
            continue
        # For new appended tx: ensure Suspense items are unconfirmed by default
        if str(tx.get('matched_ledger', '')).strip().lower() == 'suspense' and 'user_confirmed' not in tx:
            tx['user_confirmed'] = False
        merged.append(tx)
        appended_count += 1

    # Final deduplication pass to ensure no duplicate canonical keys remain (preserve first occurrence)
    seen = set()
    final_list = []
    for tx in merged:
        k = canonical_key(tx)
        if k in seen:
            continue
        seen.add(k)
        final_list.append(tx)

    statement.processed_data = final_list
    await db.commit()

    return {
        "message": "Transactions merged successfully",
        "updated": updated_count,
        "appended": appended_count,
        "final_count": len(final_list)
    }
# --- END OF ADDITION ---
# --- ADD THIS ENTIRE ENDPOINT ---

@api_router.post("/statements/{statement_id}/assign-cluster-to-ledger")
async def assign_cluster_to_ledger(
    statement_id: str,
    request: AssignClusterToLedgerRequest,
    db: AsyncSession = Depends(database.get_db)
):
    """
    Directly assigns a cluster's transactions to a ledger without creating a rule.
    Optionally adds a new ledger to known ledgers and stores samples.
    """
    statement = await db.get(models.BankStatement, statement_id)
    if not statement:
        raise HTTPException(status_code=404, detail="Statement not found")

    ledger_name = (request.ledger_name or "").strip()
    if not ledger_name:
        raise HTTPException(status_code=400, detail="Ledger name is required")

    incoming = request.transactions or []
    if not incoming:
        raise HTTPException(status_code=400, detail="No transactions provided")

    # Reuse canonical key logic from update_transactions
    def _normalize_narration(n):
        if not n:
            return ""
        s = str(n).strip().lower()
        s = re.sub(r"\s+", " ", s)
        return s

    def _normalize_date(d):
        if not d:
            return ""
        s = str(d).split(" ")[0].strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y"):
            try:
                return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
            except Exception:
                continue
        return s.lower()

    def _normalize_amount(a):
        if a is None:
            return 0.0
        try:
            s = str(a).replace(",", "").strip()
            return float(s)
        except Exception:
            try:
                return float(re.sub(r"[^\d.-]", "", str(a)))
            except Exception:
                return 0.0

    def canonical_key(tx):
        amount = (
            tx.get("Amount")
            if "Amount" in tx
            else tx.get("Amount (INR)")
            if "Amount (INR)" in tx
            else tx.get("amount")
            if "amount" in tx
            else tx.get("amt")
        )
        return f"{_normalize_narration(tx.get('Narration') or tx.get('narration'))}||{_normalize_date(tx.get('Date') or tx.get('date'))}||{_normalize_amount(amount)}"

    # Prepare incoming updates and collect samples for optional known-ledger insert
    incoming_map = {}
    samples_to_add = []

    for tx in incoming:
        clean_tx = {k: v for k, v in tx.items() if k != "_tempId"}
        clean_tx["matched_ledger"] = ledger_name
        clean_tx["user_confirmed"] = True
        incoming_map[canonical_key(clean_tx)] = clean_tx

        amount_val = tx.get("Amount") or tx.get("Amount (INR)") or tx.get("amount") or tx.get("amt") or 0
        try:
            amount = float(str(amount_val).replace(",", ""))
        except Exception:
            amount = 0.0
        cr_dr = str(tx.get("CR/DR", "")).strip().upper()
        trans_type = "Credit" if cr_dr.startswith("CR") else "Debit"
        tx_date = str(tx.get("Date", "")).split(" ")[0].strip()

        samples_to_add.append(
            {
                "narration": tx.get("Narration") or tx.get("narration") or "",
                "amount": round(amount, 2),
                "type": trans_type,
                "date": tx_date,
            }
        )

    existing = statement.processed_data or []
    merged = []
    processed_keys = set()

    # Merge incoming updates into existing list
    for orig in existing:
        key = canonical_key(orig)
        if key in incoming_map:
            merged.append({**orig, **incoming_map[key]})
            processed_keys.add(key)
        else:
            merged.append(orig)

    # Append any incoming transactions that didn't match existing keys
    for key, tx in incoming_map.items():
        if key in processed_keys:
            continue
        merged.append(tx)

    # Final deduplication pass
    seen = set()
    final_list = []
    for tx in merged:
        k = canonical_key(tx)
        if k in seen:
            continue
        seen.add(k)
        final_list.append(tx)

    statement.processed_data = final_list

    samples_added = 0
    ledger_created = False
    if request.add_to_known_ledgers:
        MAX_SAMPLES_PER_LEDGER = 25000
        query = select(models.KnownLedger).where(
            (models.KnownLedger.client_id == statement.client_id)
            & (models.KnownLedger.ledger_name == ledger_name)
        )
        result = await db.execute(query)
        db_ledger = result.scalar_one_or_none()

        existing_fingerprints = set()
        if db_ledger:
            for s in db_ledger.samples:
                s_narration, s_amount, s_type, s_date = safe_get_sample_fields(s)
                if s_narration is None or not s_narration:
                    continue
                existing_fingerprints.add(f"{s_narration}|{round(float(s_amount), 2)}|{s_type}|{s_date}")

        clean_samples = []
        for sample in samples_to_add:
            sample_date = sample.get("date", "") or ""
            clean_sample = {
                "narration": sample["narration"],
                "amount": round(float(sample["amount"]), 2),
                "type": sample["type"],
                "date": sample_date,
            }
            fingerprint = f"{clean_sample['narration']}|{clean_sample['amount']}|{clean_sample['type']}|{sample_date}"
            if fingerprint not in existing_fingerprints:
                clean_samples.append(clean_sample)
                existing_fingerprints.add(fingerprint)

        if db_ledger:
            combined_samples = db_ledger.samples + clean_samples
            if len(combined_samples) > MAX_SAMPLES_PER_LEDGER:
                combined_samples = combined_samples[-MAX_SAMPLES_PER_LEDGER:]
            db_ledger.samples = combined_samples
        else:
            db_ledger = models.KnownLedger(
                client_id=statement.client_id,
                ledger_name=ledger_name,
                samples=clean_samples[:MAX_SAMPLES_PER_LEDGER],
                is_active=True,
            )
            db.add(db_ledger)
            ledger_created = True
        samples_added = len(clean_samples)

    await db.commit()

    return {
        "message": "Cluster assigned successfully",
        "assigned": len(incoming_map),
        "final_count": len(final_list),
        "samples_added": samples_added,
        "ledger_created": ledger_created,
    }

# --- ADD THIS ENTIRE HELPER FUNCTION ---
def generate_tally_rows(vouchers: List[Dict], bank_ledger_name: str, voucher_type: str) -> List[List[Any]]:
    """
    Generates a list of rows in the two-line Tally format for Excel export.
    """
    output_rows = []
    voucher_number = 1

    for t in vouchers:
        is_contra = voucher_type == 'Contra'
        
        # Safely parse amount from standardized data
        amount = float(str(t.get('Amount', '0')).replace(',', ''))
        
        # Format date from standardized data
        date_str = t.get('Date', '').split(' ')[0]
        try:
            formatted_date = datetime.strptime(date_str, '%d/%m/%Y').strftime('%d-%b-%Y')
        except ValueError:
            formatted_date = date_str

        narration = t.get('Narration', '')
        party_ledger = t.get('matched_ledger', 'Suspense')

        # Line 1: The Party Ledger Entry
        party_cr_dr = 'Dr' if is_contra else ('Cr' if voucher_type == 'Receipt' else 'Dr')
        row1 = [
            formatted_date, voucher_type, voucher_number, '', '', 
            party_ledger, f"{amount:.2f}", party_cr_dr,
            '', '', '', '', '', '', narration
        ]
        
        # Line 2: The Bank Ledger Entry
        bank_cr_dr = 'Cr' if is_contra else ('Dr' if voucher_type == 'Receipt' else 'Cr')
        row2 = [
            '', '', '', '', '', 
            bank_ledger_name, f"{amount:.2f}", bank_cr_dr,
            '', '', '', '', '', '', ''
        ]
        
        output_rows.append(row1)
        output_rows.append(row2)
        voucher_number += 1
        
    return output_rows
# --- END OF ADDITION ---

# In server.py

# --- FIND AND REPLACE the entire generate_vouchers function WITH THESE TWO ENDPOINTS ---

@api_router.get("/vouchers/{statement_id}/summary")
async def get_voucher_summary(statement_id: str, db: AsyncSession = Depends(database.get_db)):
    """Gets the counts of each voucher type to populate the download modal."""
    
    # Eagerly load the related bank_account and client to prevent session errors.
    query = (
        select(models.BankStatement)
        .options(
            selectinload(models.BankStatement.bank_account),
            selectinload(models.BankStatement.client)
        )
        .where(models.BankStatement.id == statement_id)
    )
    result = await db.execute(query)
    statement = result.scalar_one_or_none()
    # --- END OF FIX ---

    if not statement: raise HTTPException(status_code=404, detail="Statement not found")
    
    # Now we can safely access the related objects because they were pre-loaded.
    bank_account = statement.bank_account
    client = statement.client

    if not bank_account: raise HTTPException(status_code=404, detail="Bank account not found")
    if not client: raise HTTPException(status_code=404, detail="Client not found")
    print("\n" + "="*50)
    print("INSIDE GET_VOUCHER_SUMMARY")
    print(f"  - Statement ID: {statement.id}")
    print(f"  - Raw processed_data from DB: {statement.processed_data}")
    
    processed_data = statement.processed_data or []
    
    print(f"  - Length of processed_data list being used: {len(processed_data)}")
    print("="*50 + "\n")
    # --- END: TEMPORARY DIAGNOSTIC CODE ---
    contra_list = set(bank_account.contra_list or [])
    filter_list = set(bank_account.filter_list or []) # <-- 1. Load the filter list

    receipts = 0
    payments = 0
    contras = 0

    for t in processed_data:
        cr_dr_val = str(t.get("CR/DR", "")).strip().upper()
        ledger = t.get("matched_ledger")
        if ledger in filter_list:
            continue

        if cr_dr_val.startswith("CR"):
            receipts += 1
        elif cr_dr_val.startswith("DR"):
            if ledger in contra_list:
                contras += 1
            else:
                payments += 1
    
    first_date_str = next((t.get("Date") for t in processed_data if t.get("Date")), None)
    month_year = "vouchers"
    if first_date_str:
        try:
            month_year = datetime.strptime(first_date_str.split(' ')[0], '%d/%m/%Y').strftime("%B_%Y")
        except ValueError:
            pass
            
    sanitized_ledger_name = bank_account.ledger_name.replace("/", "_").replace("\\", "_")
    suggested_filename = f"{client.name}_{sanitized_ledger_name}_{month_year}"

    return {
        "receipt_count": receipts,
        "payment_count": payments,
        "contra_count": contras,
        "suggested_filename": suggested_filename
    }

@api_router.post("/generate-vouchers/{statement_id}")
async def generate_vouchers(statement_id: str, request: VoucherGenerationRequest, db: AsyncSession = Depends(database.get_db)):
    """Generates a Tally-compatible XLSX file by populating a template."""
    statement = await db.get(models.BankStatement, statement_id)
    if not statement: raise HTTPException(status_code=404, detail="Statement not found")
    bank_account = await db.get(models.BankAccount, statement.bank_account_id)
    if not bank_account: raise HTTPException(status_code=404, detail="Bank account not found")

    # Check data integrity before generating vouchers
    integrity_check = await check_data_integrity(statement_id, db)
    if not integrity_check.is_match:
        raise HTTPException(
            status_code=400,
            detail=f"Data integrity check failed: {integrity_check.missing_count} transaction(s) are missing from processed data. Please rebuild missing transactions before generating vouchers."
        )

    processed_data = statement.processed_data or []
    contra_list = set(bank_account.contra_list or [])
    filter_list = set(bank_account.filter_list or [])
    bank_ledger_name = bank_account.ledger_name
    
    # Sort data into lists
    receipt_data, payment_data, contra_data = [], [], []
    for t in processed_data:
        cr_dr_val = str(t.get("CR/DR", "")).strip().upper()
        ledger = t.get("matched_ledger")

        if ledger in filter_list:
            continue
        if cr_dr_val.startswith("CR"):
            receipt_data.append(t)
        elif cr_dr_val.startswith("DR"):
            if ledger in contra_list:
                contra_data.append(t)
            else:
                payment_data.append(t)

    # Load the template workbook
    template_path = "templates/AccountingVouchers.xlsx"
    try:
        workbook = openpyxl.load_workbook(template_path)
        template_sheet = workbook["Accounting Voucher"]
    except (FileNotFoundError, KeyError):
        raise HTTPException(status_code=500, detail="Tally template file not found or is missing 'Accounting Voucher' sheet.")

    # Process each selected voucher type
    if request.include_receipts and receipt_data:
        receipt_sheet = workbook.copy_worksheet(template_sheet)
        receipt_sheet.title = "Receipts"
        for row in generate_tally_rows(receipt_data, bank_ledger_name, "Receipt"):
            receipt_sheet.append(row)

    if request.include_payments and payment_data:
        payment_sheet = workbook.copy_worksheet(template_sheet)
        payment_sheet.title = "Payments"
        for row in generate_tally_rows(payment_data, bank_ledger_name, "Payment"):
            payment_sheet.append(row)

    if request.include_contras and contra_data:
        contra_sheet = workbook.copy_worksheet(template_sheet)
        contra_sheet.title = "Contras"
        for row in generate_tally_rows(contra_data, bank_ledger_name, "Contra"):
            contra_sheet.append(row)
    
    # Remove the original template sheet
    del workbook["Accounting Voucher"]

    # Move the "Read Me" sheet to the very end, if it exists
    if "Accounting Voucher (Read Me)" in workbook.sheetnames:
        read_me_sheet = workbook["Accounting Voucher (Read Me)"]
        workbook.move_sheet(read_me_sheet, offset=len(workbook.sheetnames))
    
    # Save to an in-memory buffer
    output_buffer = io.BytesIO()
    workbook.save(output_buffer)
    output_buffer.seek(0)
    
    final_filename = f"{request.filename}.xlsx"
    headers = {'Content-Disposition': f'attachment; filename="{final_filename}"'}
    return StreamingResponse(output_buffer, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
# --- END OF REPLACEMENT ---

@api_router.post("/ai-improve-regex")
async def ai_improve_regex(request: AIRegexRequest):
    """Use AI to improve regex patterns"""
    improved_regex = await get_ai_improved_regex(
        request.narrations,
        request.existing_regex,
        request.use_local_llm
    )
    
    return {"improved_regex": improved_regex}

# Mock Tally API endpoints
@api_router.post("/tally/create-ledgers")
async def create_tally_ledgers(ledger_names: List[str]):
    """Mock endpoint for creating ledgers in Tally"""
    # This would integrate with actual Tally API
    return {
        "message": f"Successfully created {len(ledger_names)} ledgers in Tally",
        "ledgers": ledger_names,
        "status": "success"
    }

# Bank Account Management
@api_router.post("/bank-accounts", response_model=BankAccountModel)
async def create_bank_account(account_data: BankAccountCreate, db: AsyncSession = Depends(database.get_db)):
    """Create a new bank account for a client"""
    # Check if the client exists
    client = await db.get(models.Client, account_data.client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
        
    db_account = models.BankAccount(**account_data.dict())
    db.add(db_account)
    await db.commit()
    await db.refresh(db_account)
    return db_account

@api_router.get("/clients/{client_id}/bank-accounts", response_model=List[BankAccountModel])
async def get_client_bank_accounts(client_id: str, db: AsyncSession = Depends(database.get_db)):
    """Get all bank accounts for a specific client"""
    query = select(models.BankAccount).where(models.BankAccount.client_id == client_id)
    result = await db.execute(query)
    accounts = result.scalars().all()
    return accounts

# --- START OF ADDITION (1 of 2): The UPDATE endpoint ---
@api_router.put("/bank-accounts/{account_id}", response_model=BankAccountModel)
async def update_bank_account(
    account_id: str, 
    account_data: BankAccountUpdate, 
    db: AsyncSession = Depends(database.get_db)
):
    """Update an existing bank account."""
    db_account = await db.get(models.BankAccount, account_id)
    if not db_account:
        raise HTTPException(status_code=404, detail="Bank account not found")

    # Get the incoming data as a dictionary
    update_data = account_data.dict(exclude_unset=True)
    
    # Update the model's attributes with the new data
    for key, value in update_data.items():
        setattr(db_account, key, value)
        
    await db.commit()
    await db.refresh(db_account)
    return db_account
# --- END OF ADDITION (1 of 2) ---

@api_router.get("/bank-accounts/{account_id}", response_model=BankAccountModel)
async def get_bank_account(account_id: str, db: AsyncSession = Depends(database.get_db)):
    """Get a single bank account by its ID."""
    db_account = await db.get(models.BankAccount, account_id)
    if not db_account:
        raise HTTPException(status_code=404, detail="Bank account not found")
    return db_account

# --- START OF ADDITION (2 of 2): The DELETE endpoint with safety check ---
@api_router.delete("/bank-accounts/{account_id}", status_code=204)
async def delete_bank_account(account_id: str, db: AsyncSession = Depends(database.get_db)):
    """Delete a bank account after ensuring it has no linked statements."""
    
    # CRITICAL: Check for linked statements before deleting
    stmt_count_query = select(func.count(models.BankStatement.id)).where(
        models.BankStatement.bank_account_id == account_id
    )
    linked_statements_count = await db.scalar(stmt_count_query)
    
    if linked_statements_count > 0:
        raise HTTPException(
            status_code=400, # Bad Request, because the action is invalid
            detail=f"Cannot delete account. It is linked to {linked_statements_count} statement(s)."
        )

    # If check passes, proceed with deletion
    db_account = await db.get(models.BankAccount, account_id)
    if not db_account:
        raise HTTPException(status_code=404, detail="Bank account not found")
        
    await db.delete(db_account)
    await db.commit()
    return None # Return None for a 204 No Content response
# --- END OF ADDITION (2 of 2) ---


# --- FIND AND REPLACE THE ENTIRE get_client_statements FUNCTION ---
@api_router.get("/clients/{client_id}/statements", response_model=List[StatementMetadata])
async def get_client_statements(client_id: str, db: AsyncSession = Depends(database.get_db)):
    """Get all statement metadata for a specific client, including smart status."""
    
    query = (
        select(models.BankStatement, models.BankAccount.ledger_name)
        .join(models.BankAccount, models.BankStatement.bank_account_id == models.BankAccount.id)
        .where(models.BankStatement.client_id == client_id)
        .order_by(models.BankStatement.upload_date.desc())
    )
    result = await db.execute(query)
    
    metadata_list = []
    for stmt, bank_ledger_name in result.all():
        total = len(stmt.raw_data) if isinstance(stmt.raw_data, list) else 0
        
        # --- START: NEW CALCULATION LOGIC ---
        pending_count = 0
        matched_count = 0
        if isinstance(stmt.processed_data, list):
            for t in stmt.processed_data:
                # Count items that are 'Suspense' AND not confirmed by the user
                if t.get("matched_ledger") == "Suspense" and not t.get("user_confirmed"):
                    pending_count += 1
                # Count items that are not 'Suspense'
                if t.get("matched_ledger") != "Suspense":
                    matched_count += 1
        
        status = "Completed" if pending_count == 0 else "Needs Review"
        completion_percentage = 0.0
        if total > 0:
            completion_percentage = round(((total - pending_count) / total) * 100, 2)
        # --- END: NEW CALCULATION LOGIC ---
        # --- THIS IS THE RESTORED LOGIC ---
        statement_period = None
        if stmt.raw_data and stmt.column_mapping.get("date_column"):
            date_column = stmt.column_mapping["date_column"]
            dates = [
                pd.to_datetime(t.get(date_column), dayfirst=True, errors='coerce') 
                for t in stmt.raw_data if t.get(date_column)
            ]
            valid_dates = [d for d in dates if pd.notna(d)]
            if valid_dates:
                min_date = min(valid_dates)
                max_date = max(valid_dates)
                statement_period = f"{min_date.strftime('%d %b %Y')} - {max_date.strftime('%d %b %Y')}"
        # --- END OF RESTORED LOGIC ---

        metadata_list.append(StatementMetadata(
            id=stmt.id,
            filename=stmt.filename,
            upload_date=stmt.upload_date,
            total_transactions=total,
            matched_transactions=matched_count, # Use the calculated matched count
            bank_ledger_name=bank_ledger_name,
            statement_period=statement_period,
            status=status, # Add new status
            completion_percentage=completion_percentage # Add new percentage
        ))
    return metadata_list
# --- END OF REPLACEMENT ---

@api_router.delete("/statements/{statement_id}", status_code=204)
async def delete_statement(statement_id: str, db: AsyncSession = Depends(database.get_db)):
    """Delete a specific bank statement by its ID."""
    statement = await db.get(models.BankStatement, statement_id)
    if not statement:
        raise HTTPException(status_code=404, detail="Statement not found")
    
    await db.delete(statement)
    await db.commit()
    return None # Return None for 204 No Content response
# --- END OF ADDITION ---

@api_router.post("/tally/generate-vouchers")
async def generate_tally_vouchers(voucher_data: TallyVoucherData):
    """Generate Tally import files (Receipt, Payment, Contra vouchers)"""
    try:
        transactions = voucher_data.transactions
        bank_ledger = voucher_data.bank_ledger_name
        
        receipt_vouchers = []
        payment_vouchers = []
        contra_vouchers = []
        
        for transaction in transactions:
            ledger = transaction.get("matched_ledger", "Suspense")
            amount = float(transaction.get("amount", 0))
            narration = transaction.get("narration", "")
            date = transaction.get("date", "")
            
            # Skip excluded ledgers
            if ledger in voucher_data.excluded_ledgers:
                continue
            
            # Determine voucher type based on amount and transaction type
            if amount > 0:  # Credit to bank (Receipt)
                receipt_vouchers.append({
                    "Date": date,
                    "Voucher Type": "Receipt",
                    "Ledger Name": ledger,
                    "Amount": amount,
                    "Narration": narration,
                    "Bank Ledger": bank_ledger
                })
            else:  # Debit from bank (Payment)
                payment_vouchers.append({
                    "Date": date,
                    "Voucher Type": "Payment", 
                    "Ledger Name": ledger,
                    "Amount": abs(amount),
                    "Narration": narration,
                    "Bank Ledger": bank_ledger
                })
        
        return {
            "receipt_vouchers": receipt_vouchers,
            "payment_vouchers": payment_vouchers,
            "contra_vouchers": contra_vouchers,
            "summary": {
                "total_receipts": len(receipt_vouchers),
                "total_payments": len(payment_vouchers),
                "total_contras": len(contra_vouchers)
            }
        }
    
    except Exception as e:
        logging.error(f"Voucher generation error: {e}")
        raise HTTPException(status_code=500, detail=f"Error generating vouchers: {str(e)}")

# ADD this to server.py
@api_router.get("/statements/{statement_id}", response_model=BankStatement)
async def get_statement(statement_id: str, db: AsyncSession = Depends(database.get_db)):
    statement = await db.get(models.BankStatement, statement_id)
    if not statement:
        raise HTTPException(status_code=404, detail="Statement not found")
    return statement


# Data Integrity Check Endpoint
class PartialDataLossItem(BaseModel):
    processed_tx: Dict[str, Any]
    raw_tx: Dict[str, Any]
    missing_fields: List[str]

class DataIntegrityCheckResponse(BaseModel):
    raw_data_count: int
    processed_data_count: int
    is_match: bool
    
    # Missing rows (completely absent)
    missing_rows_count: int
    missing_rows: List[Dict[str, Any]]
    
    # Partial data loss (row exists but fields missing)
    partial_data_loss_count: int
    partial_data_loss: List[PartialDataLossItem]
    
    # Extra transactions (in processed_data but not in raw_data)
    extra_transactions_count: int
    extra_transactions: List[Dict[str, Any]]
    
    # Legacy fields for backward compatibility
    missing_count: int = 0
    missing_transactions: List[Dict[str, Any]] = []

def identify_missing_fields(processed_tx: Dict, raw_tx: Dict, mapping: Dict) -> List[str]:
    """Returns list of field names that are missing/null in processed_tx but present in raw_tx"""
    missing = []
    
    # Map processed field names to raw field names
    narration_col = mapping.get("narration_column")
    date_col = mapping.get("date_column")
    amount_col = mapping.get("amount_column")
    crdr_col = mapping.get("crdr_column")
    
    # Check Narration
    processed_narration = processed_tx.get("Narration") or processed_tx.get("narration")
    raw_narration = raw_tx.get(narration_col) if narration_col else raw_tx.get("Narration") or raw_tx.get("narration")
    if (not processed_narration or processed_narration == "") and raw_narration:
        missing.append("Narration")
    
    # Check Date
    processed_date = processed_tx.get("Date") or processed_tx.get("date")
    raw_date = raw_tx.get(date_col) if date_col else raw_tx.get("Date") or raw_tx.get("date")
    if (not processed_date or processed_date == "") and raw_date:
        missing.append("Date")
    
    # Check Amount
    processed_amount = processed_tx.get("Amount") or processed_tx.get("Amount (INR)") or processed_tx.get("amount") or processed_tx.get("amt")
    raw_amount = raw_tx.get(amount_col) if amount_col else raw_tx.get("Amount") or raw_tx.get("Amount (INR)") or raw_tx.get("amount") or raw_tx.get("amt")
    if (processed_amount is None or processed_amount == "" or processed_amount == 0) and raw_amount and raw_amount != 0:
        missing.append("Amount")
    
    # Check CR/DR
    processed_crdr = processed_tx.get("CR/DR") or processed_tx.get("crdr")
    raw_crdr = raw_tx.get(crdr_col) if crdr_col else raw_tx.get("CR/DR") or raw_tx.get("crdr")
    if (not processed_crdr or processed_crdr == "") and raw_crdr:
        missing.append("CR/DR")
    
    return missing

@api_router.get("/statements/{statement_id}/data-integrity-check", response_model=DataIntegrityCheckResponse)
async def check_data_integrity(statement_id: str, db: AsyncSession = Depends(database.get_db)):
    """Check if all raw_data transactions are present in processed_data, detecting missing rows, partial data loss, and extra transactions."""
    statement = await db.get(models.BankStatement, statement_id)
    if not statement:
        raise HTTPException(status_code=404, detail="Statement not found")
    
    raw_data = statement.raw_data or []
    processed_data = statement.processed_data or []
    
    raw_data_count = len(raw_data)
    processed_data_count = len(processed_data)
    
    # Use the same normalization functions from update_transactions
    def _normalize_narration(n):
        if not n:
            return ""
        s = str(n).strip().lower()
        s = re.sub(r'\s+', ' ', s)
        return s

    def _normalize_date(d):
        if not d:
            return ""
        s = str(d).split(' ')[0].strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y"):
            try:
                return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
            except Exception:
                continue
        return s.lower()

    def _normalize_amount(a):
        if a is None:
            return 0.0
        try:
            s = str(a).replace(',', '').strip()
            return float(s)
        except Exception:
            try:
                return float(re.sub(r'[^\d.-]', '', str(a)))
            except Exception:
                return 0.0

    def get_field_value(tx, field_name, mapping=None):
        """Helper to get field value from either raw or processed transaction"""
        if mapping:  # Raw transaction
            col = mapping.get(field_name)
            if col:
                return tx.get(col)
            # Fallback
            return tx.get(field_name) or tx.get(field_name.lower())
        else:  # Processed transaction
            return tx.get(field_name) or tx.get(field_name.lower())

    def canonical_key_from_raw(tx, mapping):
        """Create canonical key from raw_data transaction using column mapping."""
        narration = get_field_value(tx, "narration_column", mapping) or tx.get("Narration") or tx.get("narration")
        date = get_field_value(tx, "date_column", mapping) or tx.get("Date") or tx.get("date")
        amount = get_field_value(tx, "amount_column", mapping) or tx.get("Amount") or tx.get("Amount (INR)") or tx.get("amount") or tx.get("amt")
        
        return f"{_normalize_narration(narration)}||{_normalize_date(date)}||{_normalize_amount(amount)}"
    
    def canonical_key_from_processed(tx):
        """Create canonical key from processed_data transaction."""
        narration = tx.get("Narration") or tx.get("narration")
        date = tx.get("Date") or tx.get("date")
        amount = tx.get("Amount") or tx.get("Amount (INR)") or tx.get("amount") or tx.get("amt")
        
        return f"{_normalize_narration(narration)}||{_normalize_date(date)}||{_normalize_amount(amount)}"
    
    def fuzzy_key_narration_date(tx, mapping=None):
        """Create fuzzy key using narration and date only"""
        if mapping:
            narration = get_field_value(tx, "narration_column", mapping) or tx.get("Narration") or tx.get("narration")
            date = get_field_value(tx, "date_column", mapping) or tx.get("Date") or tx.get("date")
        else:
            narration = tx.get("Narration") or tx.get("narration")
            date = tx.get("Date") or tx.get("date")
        return f"{_normalize_narration(narration)}||{_normalize_date(date)}"
    
    def fuzzy_key_narration_amount(tx, mapping=None):
        """Create fuzzy key using narration and amount only"""
        if mapping:
            narration = get_field_value(tx, "narration_column", mapping) or tx.get("Narration") or tx.get("narration")
            amount = get_field_value(tx, "amount_column", mapping) or tx.get("Amount") or tx.get("Amount (INR)") or tx.get("amount") or tx.get("amt")
        else:
            narration = tx.get("Narration") or tx.get("narration")
            amount = tx.get("Amount") or tx.get("Amount (INR)") or tx.get("amount") or tx.get("amt")
        return f"{_normalize_narration(narration)}||{_normalize_amount(amount)}"
    
    def fuzzy_key_date_amount(tx, mapping=None):
        """Create fuzzy key using date and amount only"""
        if mapping:
            date = get_field_value(tx, "date_column", mapping) or tx.get("Date") or tx.get("date")
            amount = get_field_value(tx, "amount_column", mapping) or tx.get("Amount") or tx.get("Amount (INR)") or tx.get("amount") or tx.get("amt")
        else:
            date = tx.get("Date") or tx.get("date")
            amount = tx.get("Amount") or tx.get("Amount (INR)") or tx.get("amount") or tx.get("amt")
        return f"{_normalize_date(date)}||{_normalize_amount(amount)}"
    
    mapping = statement.column_mapping or {}
    
    # Build maps for matching
    processed_exact_keys = {}  # exact_key -> transaction
    processed_fuzzy_narration_date = {}  # fuzzy_key -> transaction
    processed_fuzzy_narration_amount = {}  # fuzzy_key -> transaction
    processed_fuzzy_date_amount = {}  # fuzzy_key -> transaction
    
    for tx in processed_data:
        exact_key = canonical_key_from_processed(tx)
        processed_exact_keys[exact_key] = tx
        
        # Build fuzzy keys (only if fields are present)
        if tx.get("Narration") and tx.get("Date"):
            processed_fuzzy_narration_date[fuzzy_key_narration_date(tx)] = tx
        if tx.get("Narration") and tx.get("Amount"):
            processed_fuzzy_narration_amount[fuzzy_key_narration_amount(tx)] = tx
        if tx.get("Date") and tx.get("Amount"):
            processed_fuzzy_date_amount[fuzzy_key_date_amount(tx)] = tx
    
    # Track what we've matched
    matched_raw_keys = set()
    missing_rows = []
    partial_data_loss = []
    
    # Check each raw_data transaction
    for raw_tx in raw_data:
        raw_exact_key = canonical_key_from_raw(raw_tx, mapping)
        
        # Try exact match first
        if raw_exact_key in processed_exact_keys:
            matched_raw_keys.add(raw_exact_key)
            # Check for missing fields even in exact match
            matched_processed = processed_exact_keys[raw_exact_key]
            missing_fields = identify_missing_fields(matched_processed, raw_tx, mapping)
            if missing_fields:
                partial_data_loss.append(PartialDataLossItem(
                    processed_tx=matched_processed,
                    raw_tx=raw_tx,
                    missing_fields=missing_fields
                ))
            continue
        
        # Try fuzzy matches
        matched_processed = None
        fuzzy_key_used = None
        
        # Try narration + date
        raw_fuzzy_nd = fuzzy_key_narration_date(raw_tx, mapping)
        if raw_fuzzy_nd in processed_fuzzy_narration_date:
            matched_processed = processed_fuzzy_narration_date[raw_fuzzy_nd]
            fuzzy_key_used = raw_fuzzy_nd
        # Try narration + amount
        elif fuzzy_key_narration_amount(raw_tx, mapping) in processed_fuzzy_narration_amount:
            matched_processed = processed_fuzzy_narration_amount[fuzzy_key_narration_amount(raw_tx, mapping)]
        # Try date + amount
        elif fuzzy_key_date_amount(raw_tx, mapping) in processed_fuzzy_date_amount:
            matched_processed = processed_fuzzy_date_amount[fuzzy_key_date_amount(raw_tx, mapping)]
        
        if matched_processed:
            # Found fuzzy match - this is partial data loss
            missing_fields = identify_missing_fields(matched_processed, raw_tx, mapping)
            if missing_fields:
                partial_data_loss.append(PartialDataLossItem(
                    processed_tx=matched_processed,
                    raw_tx=raw_tx,
                    missing_fields=missing_fields
                ))
            # Mark as matched
            matched_raw_keys.add(raw_exact_key)
        else:
            # No match at all - completely missing row
            missing_rows.append(raw_tx)
    
    # Find extra transactions (in processed_data but not in raw_data)
    # Track which processed transactions have been matched to raw transactions
    # This helps identify duplicates - if multiple processed transactions match the same raw transaction,
    # all but one are duplicates/extra
    
    # Build maps of raw transactions by their various keys
    raw_exact_keys = {}  # exact_key -> list of raw transactions (for tracking)
    raw_fuzzy_nd_keys = {}  # fuzzy_key -> list of raw transactions
    raw_fuzzy_na_keys = {}
    raw_fuzzy_da_keys = {}
    
    for raw_tx in raw_data:
        exact_key = canonical_key_from_raw(raw_tx, mapping)
        if exact_key not in raw_exact_keys:
            raw_exact_keys[exact_key] = []
        raw_exact_keys[exact_key].append(raw_tx)
        
        fuzzy_nd = fuzzy_key_narration_date(raw_tx, mapping)
        if fuzzy_nd not in raw_fuzzy_nd_keys:
            raw_fuzzy_nd_keys[fuzzy_nd] = []
        raw_fuzzy_nd_keys[fuzzy_nd].append(raw_tx)
        
        fuzzy_na = fuzzy_key_narration_amount(raw_tx, mapping)
        if fuzzy_na not in raw_fuzzy_na_keys:
            raw_fuzzy_na_keys[fuzzy_na] = []
        raw_fuzzy_na_keys[fuzzy_na].append(raw_tx)
        
        fuzzy_da = fuzzy_key_date_amount(raw_tx, mapping)
        if fuzzy_da not in raw_fuzzy_da_keys:
            raw_fuzzy_da_keys[fuzzy_da] = []
        raw_fuzzy_da_keys[fuzzy_da].append(raw_tx)
    
    # Track which processed transactions have been matched
    matched_processed = set()  # Set of indices of processed transactions that matched
    
    # First pass: match processed transactions to raw transactions
    # Mark exact matches
    for idx, tx in enumerate(processed_data):
        exact_key = canonical_key_from_processed(tx)
        if exact_key in raw_exact_keys:
            matched_processed.add(idx)
    
    # Second pass: mark fuzzy matches (only if not already matched)
    for idx, tx in enumerate(processed_data):
        if idx in matched_processed:
            continue  # Already matched
        
        fuzzy_nd = fuzzy_key_narration_date(tx)
        fuzzy_na = fuzzy_key_narration_amount(tx)
        fuzzy_da = fuzzy_key_date_amount(tx)
        
        # Check if any fuzzy key matches
        if (fuzzy_nd in raw_fuzzy_nd_keys or 
            fuzzy_na in raw_fuzzy_na_keys or 
            fuzzy_da in raw_fuzzy_da_keys):
            matched_processed.add(idx)
    
    # Third pass: identify duplicates
    # For each raw transaction, count how many processed transactions match it
    # If more than one processed transaction matches the same raw transaction, mark extras as duplicates
    raw_match_counts = {}  # raw_key -> count of processed transactions that matched it
    
    for idx, tx in enumerate(processed_data):
        if idx not in matched_processed:
            continue  # Not matched, will be handled as extra below
        
        exact_key = canonical_key_from_processed(tx)
        if exact_key in raw_exact_keys:
            if exact_key not in raw_match_counts:
                raw_match_counts[exact_key] = 0
            raw_match_counts[exact_key] += 1
        else:
            # Check fuzzy matches
            fuzzy_nd = fuzzy_key_narration_date(tx)
            fuzzy_na = fuzzy_key_narration_amount(tx)
            fuzzy_da = fuzzy_key_date_amount(tx)
            
            matched_key = None
            if fuzzy_nd in raw_fuzzy_nd_keys:
                matched_key = fuzzy_nd
            elif fuzzy_na in raw_fuzzy_na_keys:
                matched_key = fuzzy_na
            elif fuzzy_da in raw_fuzzy_da_keys:
                matched_key = fuzzy_da
            
            if matched_key:
                if matched_key not in raw_match_counts:
                    raw_match_counts[matched_key] = 0
                raw_match_counts[matched_key] += 1
    
    # Now identify duplicates: if multiple processed transactions match the same raw transaction,
    # keep track of which ones are duplicates
    duplicate_indices = set()
    raw_match_used = {}  # raw_key -> count of how many we've "used" (should be max 1 per raw transaction)
    
    for idx, tx in enumerate(processed_data):
        if idx not in matched_processed:
            continue
        
        exact_key = canonical_key_from_processed(tx)
        matched_key = None
        
        if exact_key in raw_exact_keys:
            matched_key = exact_key
        else:
            fuzzy_nd = fuzzy_key_narration_date(tx)
            fuzzy_na = fuzzy_key_narration_amount(tx)
            fuzzy_da = fuzzy_key_date_amount(tx)
            
            if fuzzy_nd in raw_fuzzy_nd_keys:
                matched_key = fuzzy_nd
            elif fuzzy_na in raw_fuzzy_na_keys:
                matched_key = fuzzy_na
            elif fuzzy_da in raw_fuzzy_da_keys:
                matched_key = fuzzy_da
        
        if matched_key:
            if matched_key not in raw_match_used:
                raw_match_used[matched_key] = 0
            raw_match_used[matched_key] += 1
            
            # If this is the second+ match for this raw transaction, it's a duplicate
            if raw_match_used[matched_key] > 1:
                duplicate_indices.add(idx)
    
    # Extra transactions are: unmatched + duplicates
    extra_transactions = []
    for idx, tx in enumerate(processed_data):
        if idx not in matched_processed or idx in duplicate_indices:
            extra_transactions.append(tx)
    
    missing_rows_count = len(missing_rows)
    partial_data_loss_count = len(partial_data_loss)
    extra_transactions_count = len(extra_transactions)
    
    is_match = (missing_rows_count == 0 and 
                partial_data_loss_count == 0 and 
                extra_transactions_count == 0 and 
                raw_data_count == processed_data_count)
    
    return DataIntegrityCheckResponse(
        raw_data_count=raw_data_count,
        processed_data_count=processed_data_count,
        is_match=is_match,
        missing_rows_count=missing_rows_count,
        missing_rows=missing_rows,
        partial_data_loss_count=partial_data_loss_count,
        partial_data_loss=partial_data_loss,
        extra_transactions_count=extra_transactions_count,
        extra_transactions=extra_transactions,
        # Legacy fields
        missing_count=missing_rows_count,
        missing_transactions=missing_rows
    )


# Rebuild Missing Transactions Endpoint
class RebuildRequest(BaseModel):
    update_partial_loss: bool = True  # Whether to update rows with missing fields
    add_missing_rows: bool = True  # Whether to add completely missing rows
    remove_extra_transactions: bool = False  # Whether to remove extra transactions (optional)

class RebuildMissingTransactionsResponse(BaseModel):
    message: str
    updated_count: int  # Rows updated with missing fields
    added_count: int  # New rows added
    removed_count: int  # Extra rows removed
    final_processed_count: int
    # Legacy fields
    rebuilt_count: int = 0

@api_router.post("/statements/{statement_id}/rebuild-missing-transactions", response_model=RebuildMissingTransactionsResponse)
async def rebuild_missing_transactions(
    statement_id: str, 
    request: RebuildRequest = RebuildRequest(),
    db: AsyncSession = Depends(database.get_db)
):
    """Rebuild missing transactions from raw_data, update rows with missing fields, and optionally remove extra transactions."""
    statement = await db.get(models.BankStatement, statement_id)
    if not statement:
        raise HTTPException(status_code=404, detail="Statement not found")
    
    # First check what needs to be fixed
    integrity_check = await check_data_integrity(statement_id, db)
    
    existing = list(statement.processed_data or [])  # Make a copy to modify
    mapping = statement.column_mapping or {}
    updated_count = 0
    added_count = 0
    removed_count = 0
    
    # Helper functions for matching (reuse from integrity check)
    def _normalize_narration(n):
        if not n:
            return ""
        s = str(n).strip().lower()
        s = re.sub(r'\s+', ' ', s)
        return s

    def _normalize_date(d):
        if not d:
            return ""
        s = str(d).split(' ')[0].strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y"):
            try:
                return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
            except Exception:
                continue
        return s.lower()

    def _normalize_amount(a):
        if a is None:
            return 0.0
        try:
            s = str(a).replace(',', '').strip()
            return float(s)
        except Exception:
            try:
                return float(re.sub(r'[^\d.-]', '', str(a)))
            except Exception:
                return 0.0

    def canonical_key(tx):
        amount = tx.get("Amount") if "Amount" in tx else tx.get("Amount (INR)") if "Amount (INR)" in tx else tx.get("amount") if "amount" in tx else tx.get("amt")
        return f"{_normalize_narration(tx.get('Narration') or tx.get('narration'))}||{_normalize_date(tx.get('Date') or tx.get('date'))}||{_normalize_amount(amount)}"
    
    def get_field_from_raw(raw_tx, field_name, mapping):
        """Get field value from raw transaction using mapping"""
        if field_name == "Narration":
            col = mapping.get("narration_column")
            return raw_tx.get(col) if col else raw_tx.get("Narration") or raw_tx.get("narration")
        elif field_name == "Date":
            col = mapping.get("date_column")
            return raw_tx.get(col) if col else raw_tx.get("Date") or raw_tx.get("date")
        elif field_name == "Amount":
            col = mapping.get("amount_column")
            return raw_tx.get(col) if col else raw_tx.get("Amount") or raw_tx.get("Amount (INR)") or raw_tx.get("amount") or raw_tx.get("amt")
        elif field_name == "CR/DR":
            col = mapping.get("crdr_column")
            return raw_tx.get(col) if col else raw_tx.get("CR/DR") or raw_tx.get("crdr")
        return None
    
    # Create a map of processed transactions by canonical key for quick lookup
    processed_map = {}
    for idx, tx in enumerate(existing):
        key = canonical_key(tx)
        processed_map[key] = (idx, tx)
    
    # Step 1: Update existing rows with missing fields (partial data loss)
    if request.update_partial_loss and integrity_check.partial_data_loss:
        for partial_loss_item in integrity_check.partial_data_loss:
            processed_tx = partial_loss_item.processed_tx
            raw_tx = partial_loss_item.raw_tx
            missing_fields = partial_loss_item.missing_fields
            
            # Find the existing transaction in processed_data
            processed_key = canonical_key(processed_tx)
            if processed_key in processed_map:
                idx, existing_tx = processed_map[processed_key]
                # Update only the missing fields from raw_data
                updated_tx = existing_tx.copy()
                for field in missing_fields:
                    raw_value = get_field_from_raw(raw_tx, field, mapping)
                    if field == "Amount":
                        # Standardize amount
                        updated_tx[field] = raw_value
                    elif field == "Date":
                        # Standardize date using create_standardized_transaction logic
                        if raw_value:
                            try:
                                parsed_date = pd.to_datetime(raw_value, dayfirst=True, errors='coerce')
                                if pd.notna(parsed_date):
                                    updated_tx[field] = parsed_date.strftime('%d/%m/%Y')
                                else:
                                    updated_tx[field] = str(raw_value)
                            except:
                                updated_tx[field] = str(raw_value)
                    elif field == "CR/DR":
                        # Standardize CR/DR
                        if raw_value:
                            clean_cr_dr = str(raw_value).strip().replace(".", "").upper()
                            updated_tx[field] = clean_cr_dr
                    else:
                        updated_tx[field] = raw_value
                
                # Preserve existing matched_ledger and user_confirmed
                existing[idx] = updated_tx
                updated_count += 1
    
    # Step 2: Add completely missing rows
    if request.add_missing_rows and integrity_check.missing_rows:
        standardized_missing = []
        for raw_tx in integrity_check.missing_rows:
            standardized = create_standardized_transaction(raw_tx, mapping)
            standardized["matched_ledger"] = "Suspense"
            standardized["user_confirmed"] = False
            standardized_missing.append(standardized)
            added_count += 1
        
        # Append new transactions
        existing.extend(standardized_missing)
    
    # Step 3: Remove extra transactions (if requested)
    if request.remove_extra_transactions and integrity_check.extra_transactions:
        extra_keys = {canonical_key(tx) for tx in integrity_check.extra_transactions}
        existing = [tx for tx in existing if canonical_key(tx) not in extra_keys]
        removed_count = len(integrity_check.extra_transactions)
    
    # Final deduplication pass
    seen = set()
    final_list = []
    for tx in existing:
        k = canonical_key(tx)
        if k in seen:
            continue
        seen.add(k)
        final_list.append(tx)

    statement.processed_data = final_list
    await db.commit()
    await db.refresh(statement)
    
    # Build success message
    messages = []
    if updated_count > 0:
        messages.append(f"Updated {updated_count} row(s) with missing fields")
    if added_count > 0:
        messages.append(f"Added {added_count} missing row(s)")
    if removed_count > 0:
        messages.append(f"Removed {removed_count} extra row(s)")
    
    message = "Successfully completed: " + ", ".join(messages) if messages else "No changes needed"
    
    return RebuildMissingTransactionsResponse(
        message=message,
        updated_count=updated_count,
        added_count=added_count,
        removed_count=removed_count,
        final_processed_count=len(final_list),
        rebuilt_count=added_count  # Legacy field
    )


# Devalued Keyword Management
class KeywordCreate(BaseModel):
    keyword: str

@api_router.get("/keywords", response_model=List[str])
async def get_devalued_keywords(db: AsyncSession = Depends(database.get_db)):
    """Get all devalued keywords"""
    result = await db.execute(select(models.DevaluedKeyword.keyword).order_by(models.DevaluedKeyword.keyword))
    return result.scalars().all()

@api_router.post("/keywords")
async def add_devalued_keyword(keyword_data: KeywordCreate, db: AsyncSession = Depends(database.get_db)):
    """Add a new devalued keyword"""
    keyword = keyword_data.keyword.upper().strip()
    if not keyword:
        raise HTTPException(status_code=400, detail="Keyword cannot be empty")
    
    # Check if it already exists
    exists = await db.scalar(select(models.DevaluedKeyword).where(models.DevaluedKeyword.keyword == keyword))
    if exists:
        raise HTTPException(status_code=400, detail="Keyword already exists")

    db_keyword = models.DevaluedKeyword(keyword=keyword)
    db.add(db_keyword)
    await db.commit()
    return {"message": "Keyword added successfully", "keyword": keyword}
    
# --- END OF NEW ENDPOINT BLOCK ---
# --- ADD THIS ENTIRE BLOCK OF NEW ENDPOINTS ---

# Tally Ledger History Management
@api_router.post("/clients/{client_id}/upload-ledger-history")
async def upload_ledger_history(client_id: str, file: UploadFile = File(...), db: AsyncSession = Depends(database.get_db)):
    """
    Parses a Tally Day Book export (Excel format), extracts ledger and narration
    data, and intelligently merges it with existing known ledger data.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")

    try:
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content), header=None)

        grouped_samples = defaultdict(list)
        
        # Iterate through the DataFrame rows
        for i, row in df.iterrows():
            # A transaction starts on a row with a valid date in the first column
            if pd.notna(row.iloc[0]) and isinstance(row.iloc[0], (datetime, pd.Timestamp)):
                if i + 1 < len(df): # Ensure there is a next row for the narration
                    
                    # Extract and format date as DD/MM/YYYY
                    tx_date = ""
                    try:
                        if isinstance(row.iloc[0], pd.Timestamp):
                            tx_date = row.iloc[0].strftime('%d/%m/%Y')
                        elif isinstance(row.iloc[0], datetime):
                            tx_date = row.iloc[0].strftime('%d/%m/%Y')
                    except:
                        tx_date = ""
                    
                    ledger_name = str(row.iloc[2]).strip()
                    narration = str(df.iloc[i + 1].iloc[2]).strip()
                    
                    debit_val = pd.to_numeric(row.iloc[4], errors='coerce')
                    credit_val = pd.to_numeric(row.iloc[5], errors='coerce')

                    amount = 0.0
                    trans_type = None

                    if pd.notna(debit_val) and debit_val > 0:
                        amount = debit_val
                        trans_type = "Debit"
                    elif pd.notna(credit_val) and credit_val > 0:
                        amount = credit_val
                        trans_type = "Credit"
                    
                    if ledger_name and narration and trans_type:
                        sample = {
                            "narration": narration,
                            "amount": round(float(amount), 2),
                            "type": trans_type,
                            "date": tx_date
                        }
                        grouped_samples[ledger_name].append(sample)
        
        if not grouped_samples:
            raise HTTPException(status_code=400, detail="No valid ledger data found in the uploaded file. Please check the format.")

        # --- Database Merging Logic ---
        MAX_SAMPLES_PER_LEDGER = 25000
        new_ledgers_created = 0
        ledgers_updated = 0

        for ledger_name, new_samples in grouped_samples.items():
            # Find if this ledger already exists for the client
            query = select(models.KnownLedger).where(
                (models.KnownLedger.client_id == client_id) &
                (models.KnownLedger.ledger_name == ledger_name)
            )
            result = await db.execute(query)
            db_ledger = result.scalar_one_or_none()

            if not db_ledger:
                db_ledger = models.KnownLedger(
                    client_id=client_id,
                    ledger_name=ledger_name,
                    samples=[]
                )
                db.add(db_ledger)
                new_ledgers_created += 1
            else:
                ledgers_updated += 1
            
            # Use a set for efficient de-duplication (include date and round amounts)
            existing_fingerprints = set()
            for s in db_ledger.samples:
                s_narration, s_amount, s_type, s_date = safe_get_sample_fields(s)
                if s_narration is None or not s_narration:
                    continue
                existing_fingerprints.add(f"{s_narration}|{round(float(s_amount), 2)}|{s_type}|{s_date}")
            
            unique_new_samples = []
            for sample in new_samples:
                sample_date = sample.get('date', '') or ''
                fingerprint = f"{sample['narration']}|{round(float(sample['amount']), 2)}|{sample['type']}|{sample_date}"
                if fingerprint not in existing_fingerprints:
                    unique_new_samples.append(sample)
                    existing_fingerprints.add(fingerprint)
            
            # Combine and apply FIFO cap
            combined_samples = db_ledger.samples + unique_new_samples
            if len(combined_samples) > MAX_SAMPLES_PER_LEDGER:
                combined_samples = combined_samples[-MAX_SAMPLES_PER_LEDGER:] # Keep the newest
            
            db_ledger.samples = combined_samples
        
        await db.commit()

        return {
            "message": "Ledger history uploaded and merged successfully.",
            "ledgers_found": len(grouped_samples),
            "new_ledgers_created": new_ledgers_created,
            "existing_ledgers_updated": ledgers_updated
        }

    except Exception as e:
        await db.rollback()
        logging.error(f"Tally history upload error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@api_router.get("/clients/{client_id}/known-ledgers", response_model=List[str])
async def get_known_ledgers_for_client(client_id: str, include_inactive: bool = Query(False), db: AsyncSession = Depends(database.get_db)):
    """Gets a simple, distinct list of known ledger names for a client. By default only returns active ledgers."""
    query = select(models.KnownLedger.ledger_name).where(models.KnownLedger.client_id == client_id)
    if not include_inactive:
        query = query.where(models.KnownLedger.is_active == True)
    query = query.distinct().order_by(models.KnownLedger.ledger_name)
    result = await db.execute(query)
    return result.scalars().all()

@api_router.get("/clients/{client_id}/ledgers", response_model=PaginatedLedgersResponse)
async def get_paginated_ledgers_for_client(
    client_id: str, 
    page: int = Query(1, ge=1),
    limit: int = Query(12, ge=1, le=1000),
    search: Optional[str] = Query(None),
    include_inactive: bool = Query(False),
    db: AsyncSession = Depends(database.get_db)
):
    """Gets a paginated, searchable list of known ledgers with rule counts."""
    
    # 1. Subquery to count rules per ledger_name for the specific client
    rule_counts_subquery = (
        select(
            models.LedgerRule.ledger_name,
            func.count(models.LedgerRule.id).label("rule_count")
        )
        .where(models.LedgerRule.client_id == client_id)
        .group_by(models.LedgerRule.ledger_name)
        .subquery()
    )

    # 2. Query KnownLedger records with rule counts
    known_ledgers_query = (
        select(
            models.KnownLedger.id,
            models.KnownLedger.ledger_name,
            func.json_array_length(models.KnownLedger.samples).label("sample_count"),
            func.coalesce(rule_counts_subquery.c.rule_count, 0).label("rule_count"),
            models.KnownLedger.is_active,
            models.KnownLedger.samples
        )
        .outerjoin(
            rule_counts_subquery,
            models.KnownLedger.ledger_name == rule_counts_subquery.c.ledger_name
        )
        .where(models.KnownLedger.client_id == client_id)
    )

    # Apply active filter unless include_inactive is True
    if not include_inactive:
        known_ledgers_query = known_ledgers_query.where(models.KnownLedger.is_active == True)

    # Apply search filter if provided
    if search:
        known_ledgers_query = known_ledgers_query.where(models.KnownLedger.ledger_name.ilike(f"%{search}%"))

    # Execute KnownLedger query
    known_ledgers_result = await db.execute(known_ledgers_query)
    known_ledgers_rows = known_ledgers_result.mappings().all()

    # 3. Query rule-only ledgers (ledgers that exist in LedgerRule but not in KnownLedger)
    rule_only_query = (
        select(
            models.LedgerRule.ledger_name,
            func.count(models.LedgerRule.id).label("rule_count")
        )
        .where(
            models.LedgerRule.client_id == client_id,
            ~exists(
                select(1)
                .where(
                    models.KnownLedger.client_id == client_id,
                    models.KnownLedger.ledger_name == models.LedgerRule.ledger_name
                )
            )
        )
        .group_by(models.LedgerRule.ledger_name)
    )

    # Apply search filter to rule-only query if provided
    if search:
        rule_only_query = rule_only_query.where(models.LedgerRule.ledger_name.ilike(f"%{search}%"))

    # Execute rule-only query
    rule_only_result = await db.execute(rule_only_query)
    rule_only_rows = rule_only_result.mappings().all()

    # 4. Build combined list of ledger summaries
    ledger_summaries = []
    
    # Add KnownLedger entries
    for row in known_ledgers_rows:
        last_txn_date = compute_last_transaction_date(row.get("samples"))
        ledger_summaries.append(
            KnownLedgerSummaryWithRuleCount(
                id=row["id"],
                ledger_name=row["ledger_name"],
                sample_count=row["sample_count"],
                rule_count=row["rule_count"],
                is_active=row["is_active"],
                last_transaction_date=last_txn_date,
            )
        )
    
    # Add rule-only entries (ledgers with rules but no KnownLedger record)
    for row in rule_only_rows:
        ledger_name = row["ledger_name"]
        rule_count = row["rule_count"]
        # Generate temporary ID for rule-only ledgers
        temp_id = f"rule_{ledger_name}"
        
        ledger_summaries.append(
            KnownLedgerSummaryWithRuleCount(
                id=temp_id,
                ledger_name=ledger_name,
                sample_count=0,  # No samples if no KnownLedger
                rule_count=rule_count,
                is_active=True,  # Default to active
                last_transaction_date=None,
            )
        )

    # 5. Sort by ledger_name
    ledger_summaries.sort(key=lambda x: x.ledger_name.lower())

    # 6. Apply pagination
    total_ledgers = len(ledger_summaries)
    total_pages = (total_ledgers + limit - 1) // limit if total_ledgers > 0 else 1
    start_idx = (page - 1) * limit
    end_idx = start_idx + limit
    paginated_ledgers = ledger_summaries[start_idx:end_idx]

    return PaginatedLedgersResponse(
        total_ledgers=total_ledgers,
        total_pages=total_pages,
        ledgers=paginated_ledgers,
    )



@api_router.get("/clients/{client_id}/known-ledgers/summary", response_model=List[KnownLedgerSummary])
async def get_known_ledgers_summary(client_id: str, include_inactive: bool = Query(False), db: AsyncSession = Depends(database.get_db)):
    """Gets a summary of all known ledgers for a client, including their sample counts."""
    query = select(
        models.KnownLedger.id,
        models.KnownLedger.ledger_name,
        func.json_array_length(models.KnownLedger.samples).label("sample_count"),
        models.KnownLedger.is_active
    ).where(models.KnownLedger.client_id == client_id)
    
    if not include_inactive:
        query = query.where(models.KnownLedger.is_active == True)
    
    query = query.order_by(models.KnownLedger.ledger_name)
    
    result = await db.execute(query)
    # Use .mappings() to get dict-like rows
    return [KnownLedgerSummary(**row) for row in result.mappings().all()]


@api_router.get("/known-ledgers/{ledger_id}/samples", response_model=PaginatedSamplesResponse)
async def get_ledger_samples(ledger_id: str, page: int = Query(1, ge=1), limit: int = Query(100, ge=1, le=500), db: AsyncSession = Depends(database.get_db)):
    """Gets a paginated list of samples for a specific known ledger."""
    db_ledger = await db.get(models.KnownLedger, ledger_id)
    if not db_ledger:
        raise HTTPException(status_code=404, detail="Ledger not found")

    all_samples = db_ledger.samples
    total_samples = len(all_samples)
    
    # Normalize samples to handle malformed data
    normalized_samples = []
    for s in all_samples:
        s_narration, s_amount, s_type, s_date = safe_get_sample_fields(s)
        # Skip malformed samples that can't be parsed
        if s_narration is None or not s_narration:
            continue
        normalized_samples.append({
            "narration": s_narration,
            "amount": s_amount,
            "type": s_type,
            "date": s_date
        })
    
    # Update total_samples to reflect only valid samples
    total_samples = len(normalized_samples)
    
    # Paginate in Python
    start_index = (page - 1) * limit
    end_index = start_index + limit
    paginated_samples = normalized_samples[start_index:end_index]

    return PaginatedSamplesResponse(
        total_samples=total_samples,
        samples=paginated_samples
    )

@api_router.delete("/known-ledgers/{ledger_id}/samples")
async def delete_ledger_sample(ledger_id: str, request: DeleteSampleRequest, db: AsyncSession = Depends(database.get_db)):
    """Deletes a specific sample from a known ledger's sample list."""
    db_ledger = await db.get(models.KnownLedger, ledger_id)
    if not db_ledger:
        raise HTTPException(status_code=404, detail="Ledger not found")

    samples = db_ledger.samples or []

    # Delete by index when provided
    if request.index is not None:
        if request.index < 0 or request.index >= len(samples):
            raise HTTPException(status_code=404, detail="Sample not found in ledger")
        samples.pop(request.index)
        db_ledger.samples = samples
        await db.commit()
        return {"message": "Sample deleted successfully."}

    # Otherwise delete by matching sample content (subset match)
    if request.sample:
        target = request.sample
        def matches(sample):
            return all(sample.get(k) == v for k, v in target.items())
        new_samples = []
        removed = False
        for s in samples:
            if not removed and matches(s):
                removed = True
                continue
            new_samples.append(s)
        if not removed:
            raise HTTPException(status_code=404, detail="Sample not found in ledger")
        db_ledger.samples = new_samples
        await db.commit()
        return {"message": "Sample deleted successfully."}

    raise HTTPException(status_code=400, detail="No sample specified for deletion")


# --- SELF-LEARNING LEDGER ENDPOINTS ---

def safe_get_sample_fields(sample):
    """Safely extracts narration, amount, and type from a sample, handling malformed data."""
    if not isinstance(sample, dict):
        return None, None, None, ''
    
    # Try to get narration with various key formats
    narration = (sample.get('narration') or 
                sample.get('Narration') or 
                sample.get('"narration') or  # Handle malformed key with quote
                '')
    
    # If narration is missing, check if it's stored as a key (malformed data)
    if not narration or narration.strip() in [':', ': ', '']:
        for key in sample.keys():
            # Check if key looks like a narration (long string, contains transaction details)
            if isinstance(key, str) and len(key) > 20 and key not in ['amount', 'type', 'date', 'narration', 'Narration', '"narration', 'Credit', 'Debit']:
                # This key is likely the narration text stored incorrectly as a key
                narration = key.strip()
                break
    
    # Try to get amount with various formats
    amount_str = (sample.get('amount') or 
                 sample.get('Amount') or 
                 '0')
    # Parse amount from strings like ": 504008, " or "504008"
    if isinstance(amount_str, str):
        # Extract number from string like ": 504008, "
        numbers = re.findall(r'[\d,]+\.?\d*', amount_str)
        if numbers:
            amount_str = numbers[0]
    try:
        amount_val = float(str(amount_str).replace(',', ''))
    except:
        amount_val = 0.0
    
    # Try to get type with various formats
    trans_type = (sample.get('type') or 
                 sample.get('Type') or 
                 '')
    # Check if "Credit" or "Debit" are keys (malformed data)
    if not trans_type:
        if 'Credit' in sample:
            trans_type = 'Credit'
        elif 'Debit' in sample:
            trans_type = 'Debit'
    
    # Get date
    s_date = sample.get('date', '') or ''
    
    return narration, amount_val, trans_type, s_date

@api_router.post("/clients/{client_id}/learn-ledgers")
async def learn_ledgers_from_statements(client_id: str, request: LearnLedgersRequest, db: AsyncSession = Depends(database.get_db)):
    """
    Streams the learning process from processed statements in real-time.
    Returns a JSON lines stream showing each transaction being processed.
    Does NOT commit to database - results must be accepted via accept-learned-ledgers endpoint.
    """
    
    async def generate_learning_stream():
        delay_seconds = request.delay_ms / 1000.0
        
        # Load existing ledgers for this client to track which are new vs existing
        existing_ledgers_query = select(models.KnownLedger).where(models.KnownLedger.client_id == client_id)
        existing_result = await db.execute(existing_ledgers_query)
        existing_ledgers = {ledger.ledger_name: ledger for ledger in existing_result.scalars().all()}
        
        # Track learning results in memory
        # Structure: {ledger_name: {is_new: bool, samples: [...], existing_sample_count: int}}
        learning_results = {}
        
        # Process each statement
        for stmt_index, statement_id in enumerate(request.statement_ids):
            statement = await db.get(models.BankStatement, statement_id)
            if not statement or statement.client_id != client_id:
                continue
            
            processed_data = statement.processed_data or []
            total_transactions = len(processed_data)
            
            # Emit statement_start event
            yield json.dumps({
                "type": "statement_start",
                "statement_id": statement_id,
                "filename": statement.filename,
                "total_transactions": total_transactions,
                "statement_index": stmt_index,
                "total_statements": len(request.statement_ids)
            }) + "\n"
            
            # Process each transaction
            for tx_index, transaction in enumerate(processed_data):
                ledger_name = transaction.get("matched_ledger", "").strip()
                narration = transaction.get("Narration", "").strip()
                
                # Skip Suspense transactions - they have no learning value
                if not ledger_name or ledger_name.lower() == "suspense" or not narration:
                    continue
                
                # Determine amount and type
                amount = 0.0
                try:
                    amount_val = transaction.get("Amount") or transaction.get("Amount (INR)") or 0
                    amount = float(str(amount_val).replace(",", ""))
                except:
                    amount = 0.0
                
                cr_dr = str(transaction.get("CR/DR", "")).strip().upper()
                trans_type = "Credit" if cr_dr.startswith("CR") else "Debit"
                
                # Extract date from transaction (format: DD/MM/YYYY)
                tx_date = ""
                raw_date = transaction.get("Date", "")
                if raw_date:
                    # Extract just the date part (before any time component)
                    tx_date = str(raw_date).split(' ')[0].strip()
                
                # Create sample object with date
                sample = {
                    "narration": narration,
                    "amount": round(amount, 2),  # Round for consistency
                    "type": trans_type,
                    "date": tx_date,
                    "user_confirmed": transaction.get("user_confirmed", False)
                }
                
                # Check if this ledger is new or existing
                is_new_ledger = ledger_name not in existing_ledgers and ledger_name not in learning_results
                
                # Initialize ledger in learning results if needed
                if ledger_name not in learning_results:
                    existing_ledger = existing_ledgers.get(ledger_name)
                    learning_results[ledger_name] = {
                        "is_new": ledger_name not in existing_ledgers,
                        "samples": [],
                        "existing_sample_count": len(existing_ledger.samples) if existing_ledger else 0,
                        "is_active": existing_ledger.is_active if existing_ledger else True
                    }
                
                # Create fingerprint for deduplication (round amount and include date)
                fingerprint = f"{narration}|{round(amount, 2)}|{trans_type}|{tx_date}"
                
                # Check if sample already exists (in existing ledger or in learning results)
                existing_fingerprints = set()
                if ledger_name in existing_ledgers:
                    for s in existing_ledgers[ledger_name].samples:
                        s_narration, s_amount, s_type, s_date = safe_get_sample_fields(s)
                        # Skip malformed samples that couldn't be parsed
                        if s_narration is None or not s_narration:
                            continue
                        existing_fingerprints.add(f"{s_narration}|{round(float(s_amount), 2)}|{s_type}|{s_date}")
                
                for s in learning_results[ledger_name]["samples"]:
                    s_narration, s_amount, s_type, s_date = safe_get_sample_fields(s)
                    # Skip malformed samples
                    if s_narration is None or not s_narration:
                        continue
                    existing_fingerprints.add(f"{s_narration}|{round(float(s_amount), 2)}|{s_type}|{s_date}")
                
                # Only add if not a duplicate
                if fingerprint not in existing_fingerprints:
                    learning_results[ledger_name]["samples"].append(sample)
                
                # Emit transaction event
                yield json.dumps({
                    "type": "transaction",
                    "narration": narration[:100] + "..." if len(narration) > 100 else narration,
                    "ledger_name": ledger_name,
                    "amount": amount,
                    "trans_type": trans_type,
                    "transaction_index": tx_index,
                    "total_transactions": total_transactions,
                    "statement_index": stmt_index
                }) + "\n"
                
                # Emit ledger update event
                total_samples = learning_results[ledger_name]["existing_sample_count"] + len(learning_results[ledger_name]["samples"])
                yield json.dumps({
                    "type": "ledger_update",
                    "ledger_name": ledger_name,
                    "new_sample_count": len(learning_results[ledger_name]["samples"]),
                    "total_samples": total_samples,
                    "is_new": learning_results[ledger_name]["is_new"]
                }) + "\n"
                
                # Delay for visualization
                if delay_seconds > 0:
                    await asyncio.sleep(delay_seconds)
        
        # Build final summary
        new_ledgers = [name for name, data in learning_results.items() if data["is_new"]]
        updated_ledgers = [name for name, data in learning_results.items() if not data["is_new"] and len(data["samples"]) > 0]
        total_new_samples = sum(len(data["samples"]) for data in learning_results.values())
        
        # Emit completion event with full results
        yield json.dumps({
            "type": "complete",
            "results": {
                "ledgers_found": len(learning_results),
                "new_ledgers": new_ledgers,
                "updated_ledgers": updated_ledgers,
                "total_new_samples": total_new_samples,
                "learning_data": {
                    name: {
                        "is_new": data["is_new"],
                        "new_sample_count": len(data["samples"]),
                        "existing_sample_count": data["existing_sample_count"],
                        "total_samples": data["existing_sample_count"] + len(data["samples"]),
                        "samples": data["samples"]  # Include samples for acceptance
                    }
                    for name, data in learning_results.items()
                }
            }
        }) + "\n"
    
    return StreamingResponse(
        generate_learning_stream(),
        media_type="application/x-ndjson",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
    )


@api_router.post("/clients/{client_id}/accept-learned-ledgers")
async def accept_learned_ledgers(client_id: str, request: AcceptLearnedLedgersRequest, db: AsyncSession = Depends(database.get_db)):
    """
    Accepts or rejects learned ledgers from a learning session.
    Accepted ledgers are committed to the database.
    Rejected ledgers (if new) are not created; (if existing) their new samples are discarded.
    """
    learning_data = request.learning_results.get("learning_data", {})
    
    accepted_count = 0
    rejected_count = 0
    samples_added = 0
    
    MAX_SAMPLES_PER_LEDGER = 25000
    
    for ledger_name in request.accepted_ledgers:
        if ledger_name not in learning_data:
            continue
        
        ledger_info = learning_data[ledger_name]
        new_samples = ledger_info.get("samples", [])
        
        if not new_samples:
            continue
        
        # Check if ledger exists
        query = select(models.KnownLedger).where(
            (models.KnownLedger.client_id == client_id) &
            (models.KnownLedger.ledger_name == ledger_name)
        )
        result = await db.execute(query)
        db_ledger = result.scalar_one_or_none()
        
        if db_ledger:
            # Merge samples (deduplicate) - use rounded amounts and include date
            existing_fingerprints = set()
            for s in db_ledger.samples:
                s_narration, s_amount, s_type, s_date = safe_get_sample_fields(s)
                if s_narration is None or not s_narration:
                    continue
                existing_fingerprints.add(f"{s_narration}|{round(float(s_amount), 2)}|{s_type}|{s_date}")
            
            unique_new_samples = []
            for sample in new_samples:
                # Include date in clean_sample, remove user_confirmed flag
                sample_date = sample.get("date", "") or ""
                clean_sample = {
                    "narration": sample["narration"],
                    "amount": round(float(sample["amount"]), 2),
                    "type": sample["type"],
                    "date": sample_date
                }
                fingerprint = f"{clean_sample['narration']}|{clean_sample['amount']}|{clean_sample['type']}|{sample_date}"
                if fingerprint not in existing_fingerprints:
                    unique_new_samples.append(clean_sample)
                    existing_fingerprints.add(fingerprint)
            
            combined_samples = db_ledger.samples + unique_new_samples
            if len(combined_samples) > MAX_SAMPLES_PER_LEDGER:
                combined_samples = combined_samples[-MAX_SAMPLES_PER_LEDGER:]
            
            db_ledger.samples = combined_samples
            samples_added += len(unique_new_samples)
        else:
            # Create new ledger - include date in samples
            clean_samples = [{
                "narration": s["narration"],
                "amount": round(float(s["amount"]), 2),
                "type": s["type"],
                "date": s.get("date", "") or ""
            } for s in new_samples]
            
            db_ledger = models.KnownLedger(
                client_id=client_id,
                ledger_name=ledger_name,
                samples=clean_samples[:MAX_SAMPLES_PER_LEDGER],
                is_active=True
            )
            db.add(db_ledger)
            samples_added += len(clean_samples[:MAX_SAMPLES_PER_LEDGER])
        
        accepted_count += 1
    
    # Rejected ledgers - nothing to do since we didn't commit them yet
    rejected_count = len(request.rejected_ledgers)
    
    await db.commit()
    
    return {
        "message": "Learning results processed successfully.",
        "accepted_ledgers": accepted_count,
        "rejected_ledgers": rejected_count,
        "samples_added": samples_added
    }


@api_router.patch("/known-ledgers/{ledger_id}/toggle-active")
async def toggle_ledger_active(ledger_id: str, request: ToggleLedgerActiveRequest, db: AsyncSession = Depends(database.get_db)):
    """Toggles the is_active status of a known ledger."""
    db_ledger = await db.get(models.KnownLedger, ledger_id)
    if not db_ledger:
        raise HTTPException(status_code=404, detail="Ledger not found")
    
    db_ledger.is_active = request.is_active
    await db.commit()
    await db.refresh(db_ledger)
    
    return {
        "id": db_ledger.id,
        "ledger_name": db_ledger.ledger_name,
        "is_active": db_ledger.is_active,
        "message": f"Ledger {'activated' if request.is_active else 'deactivated'} successfully."
    }


@api_router.delete("/known-ledgers/{ledger_id}", status_code=204)
async def delete_known_ledger(ledger_id: str, cascade_rules: bool = Query(True), db: AsyncSession = Depends(database.get_db)):
    """
    Deletes a known ledger by ID. Optionally cascades removal of ledger rules for the same client and ledger name.
    """
    db_ledger = await db.get(models.KnownLedger, ledger_id)
    if not db_ledger:
        raise HTTPException(status_code=404, detail="Ledger not found")

    if cascade_rules:
        await db.execute(
            delete(models.LedgerRule).where(
                models.LedgerRule.client_id == db_ledger.client_id,
                models.LedgerRule.ledger_name == db_ledger.ledger_name,
            )
        )

    await db.delete(db_ledger)
    await db.commit()
    return None


# --- END OF SELF-LEARNING LEDGER ENDPOINTS ---

# --- END OF NEW ENDPOINTS BLOCK ---
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)
# This entire block will only run if the APP_ENV environment variable is set to "production"
if os.environ.get("APP_ENV") == "production":
    
    # Mount the static files directory for the built React app
    app.mount("/static", StaticFiles(directory="build/static"), name="static")

    @app.get("/{catchall:path}", response_class=FileResponse)
    def read_root(catchall: str):
        """
        Catch-all route to serve the React index.html file for any non-API, non-static path.
        This is essential for the React Router to handle client-side routing in production.
        """
        return "build/index.html"
# --- END OF ADDITION ---